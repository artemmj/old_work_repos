import os

from django.conf import settings
from django.utils import timezone
from kombu import uuid
from openpyxl import load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.workbook import Workbook

from apps import app
from apps.address.models.city import City


@app.task
def cities_export_excel(host: str) -> str:  # noqa: WPS210
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 30

    cells = []
    coll_names = ['Город', 'Цена осмотра']

    for coll_name in coll_names:
        cell = WriteOnlyCell(ws, value=coll_name)
        cell.font = Font(bold=True)
        cells.append(cell)
    ws.append(cells)

    for city in City.objects.all():
        user_cells = [
            WriteOnlyCell(ws, value=city.title),
            WriteOnlyCell(ws, value=city.inspection_price),
        ]
        ws.append(user_cells)

    dir_path = f'/media/cities_export_excel/{timezone.now().strftime("%Y/%m/%d")}'  # noqa: WPS237
    filename = f'export_{uuid()}'
    if not os.path.isdir(dir_path):
        os.makedirs(dir_path)
    filepath = f'{dir_path}/{filename}.xlsx'
    wb.save(filepath)
    return f'{host}{filepath}'


@app.task
def cities_import_excel(file: str):  # noqa: WPS210
    ws = load_workbook(f'{settings.MEDIA_ROOT}/{file}').active
    row_pointer = 2
    col_pointer = 1

    while True:
        if not ws.cell(row=row_pointer, column=col_pointer).value:
            os.remove(f'{settings.MEDIA_ROOT}/{file}')
            break
        city_title = ws.cell(row=row_pointer, column=col_pointer).value
        price = ws.cell(row=row_pointer, column=col_pointer + 1).value
        City.objects.update_or_create(title=city_title, defaults={'inspection_price': price})
        row_pointer += 1
