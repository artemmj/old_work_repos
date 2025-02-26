from django.contrib.gis.geos import Point
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from apps.address.models import City
from apps.regions.models import Region

CITIES_FILE_PATH = 'apps/regions/management/cities_coords.xlsx'


class Command(BaseCommand):
    help = 'Импорт xlsx файла городов.'

    def handle(self, *args, **kwargs):  # noqa: WPS110
        ws = self._get_ws()
        for row_counter in range(2, ws.max_row + 1):
            # Начинаем со второй строки
            self._write_region_and_city(ws, row_counter)
        return 'Импорт успешно завершен.'

    def _get_ws(self) -> Worksheet:
        """Открывает файл, возвращает лист."""
        wb = load_workbook(CITIES_FILE_PATH, read_only=True)
        return wb.active

    def _write_region_and_city(self, ws, row_counter) -> City:
        """Считывает в каждой строке данные по городу, сохраняет/обновляет."""
        city_name = ws.cell(row_counter, 1).value
        region_name = ws.cell(row_counter, 4).value
        latitude = ws.cell(row_counter, 5).value
        longitude = ws.cell(row_counter, 6).value

        region, _ = Region.objects.get_or_create(name=region_name)
        City.objects.update_or_create(
            name=city_name,
            defaults={
                'location': Point(float(longitude), float(latitude)),
                'region': region,
            },
        )
