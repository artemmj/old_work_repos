import os
import pytest

from openpyxl import load_workbook

from apps.file.models import Image
from apps.stories.services import StoriesViewsReportService


@pytest.mark.django_db
def test_stories_views_report(
    users_factory,
    departments_factory,
    image,
    news_factory,
    stories_factory,
    stories_read_factory,
):
    departments = departments_factory(count=10, name=(f'd{_}' for _ in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]))
    users = users_factory(count=10, department=(_dep for _dep in departments))
    picture = Image.objects.first()
    news = news_factory(count=10, preview_standard=picture)
    stories = stories_factory(
        count=10,
        news=(_news for _news in news),
        preview=picture,
    )
    stories_read_factory(
        count=10,
        stories=(_stories for _stories in stories),
        user=(_user for _user in users),
    )
    workbook = StoriesViewsReportService(start='2024-01-01', end='2034-01-01').process()
    ws = workbook.active
    assert ws.cell(row=1, column=1).value == 'Сторис'
    assert ws.cell(row=2, column=2).value == 1
    assert ws.cell(row=3, column=3).value == 1
