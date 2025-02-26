import logging
from decimal import Decimal
from secrets import token_urlsafe

from django.conf import settings
from django.template.loader import render_to_string
from openpyxl import load_workbook
from xhtml2pdf import pisa

from api.v1.product.services import MergeProductTitleAndTitleAttrsService
from api.v1.reports.services.base_service import BaseReportService
from api.v1.reports.services.helpers.check_reports_directory import check_reports_directory
from apps import app
from apps.document.models import Document, DocumentStatusChoices
from apps.product.models import Product

logger = logging.getLogger('django')


@app.task
def generate_differences_report_task(serializer_data: dict, endpoint_prefx: str):
    return DifferencesReportService(serializer_data, endpoint_prefx).process()


class DifferencesReportService(BaseReportService):  # noqa: WPS338
    """Сервис для генерации отчета Ре Трейдинг счислительная (Отчет по расхождениям)."""

    def __init__(self, serializer_data: dict, endpoint_pref: str):
        super().__init__(serializer_data, endpoint_pref)
        self.is_excel = serializer_data.get('excel')

    def _create_context(self):  # noqa: WPS231
        source_data = {}
        for product in Product.objects.filter(project=self.project, remainder__isnull=False):
            source_data[product] = {
                'barcode': product.barcode,
                'title': MergeProductTitleAndTitleAttrsService(
                    project_id=product.project.id,
                    product_id=product.id,
                    product_title=product.title,
                ).process(),
                'vendor_code': product.vendor_code,
                'price': float(product.price),
                'remainder': product.remainder,
                'fact': 0,
            }

        for document in Document.objects.filter(zone__project=self.project, status=DocumentStatusChoices.READY):
            for scan_prod in document.counter_scan_task.scanned_products.all():
                if scan_prod.product not in source_data:
                    source_data[scan_prod.product]['fact'] = scan_prod.amount
                else:
                    source_data[scan_prod.product]['fact'] += scan_prod.amount

        context = {
            'products': [],
            'alls_surplus_count': 0,
            'alls_surplus_sum': 0,
            'alls_shortage_count': 0,
            'alls_shortage_sum': 0,
        }

        for prodkey in source_data.keys():
            if source_data[prodkey]['remainder'] == source_data[prodkey]['fact']:
                continue

            row = {
                'vendor_code': source_data[prodkey]['vendor_code'],
                'title': source_data[prodkey]['title'],
                'barcode': source_data[prodkey]['barcode'],
                'price': source_data[prodkey]['price'],
                'uchet': source_data[prodkey]['remainder'],
                'fact': source_data[prodkey]['fact'],
                'surplus_count': 0,
                'surplus_sum': 0,
                'shortage_count': 0,
                'shortage_sum': 0,
            }

            if source_data[prodkey]['remainder'] < 0:
                difference = abs(source_data[prodkey]['remainder'] - source_data[prodkey]['fact'])
            else:
                difference = source_data[prodkey]['fact'] - source_data[prodkey]['remainder']

            if difference > 0:
                # Излишек
                surplus_sum = difference * Decimal(source_data[prodkey]['price'])
                row['surplus_count'] = difference
                row['surplus_sum'] = difference * Decimal(source_data[prodkey]['price'])
                context['alls_surplus_count'] += difference
                context['alls_surplus_sum'] += surplus_sum
            elif difference < 0:
                # Недостача
                shortage_sum = difference * Decimal(source_data[prodkey]['price'])
                row['shortage_count'] = abs(difference)
                row['shortage_sum'] = abs(difference * Decimal(source_data[prodkey]['price']))
                context['alls_shortage_count'] += abs(difference)
                context['alls_shortage_sum'] += abs(shortage_sum)

            context['products'].append(row)

        for product in context['products']:
            product['surplus_count'] = round(product['surplus_count'], 3)
            product['surplus_sum'] = round(product['surplus_sum'], 2)
            product['shortage_count'] = round(product['shortage_count'], 3)
            product['shortage_sum'] = round(product['shortage_sum'], 2)
        context['alls_surplus_count'] = round(context['alls_surplus_count'], 3)
        context['alls_surplus_sum'] = round(context['alls_surplus_sum'], 2)
        context['alls_shortage_count'] = round(context['alls_shortage_count'], 3)
        context['alls_shortage_sum'] = round(context['alls_shortage_sum'], 2)
        return context

    def _generate_excel(self, context):  # noqa: WPS213
        worksheet = self.workbook.worksheets[0]

        alls_surplus_count = 0
        alls_surplus_sum = 0
        alls_shortage_count = 0
        alls_shortage_sum = 0

        row_pointer = 3
        for idx, product in enumerate(context.get('products'), start=1):
            worksheet.cell(row=row_pointer, column=1, value=idx)
            worksheet.cell(row=row_pointer, column=2, value=product['vendor_code'])
            worksheet.cell(row=row_pointer, column=3, value=product['title'])
            worksheet.cell(row=row_pointer, column=4, value=product['barcode'])

            difference = product['fact'] - product['uchet']
            if difference > 0:
                # Излишек
                worksheet.cell(row=row_pointer, column=7, value=0)
                worksheet.cell(row=row_pointer, column=8, value=0)
                worksheet.cell(row=row_pointer, column=5, value=difference)
                surplus_sum = difference * Decimal(product['price'])
                worksheet.cell(row=row_pointer, column=6, value=surplus_sum)
                alls_surplus_count += difference
                alls_surplus_sum += surplus_sum
            elif difference < 0:
                # Недостача
                worksheet.cell(row=row_pointer, column=5, value=0)
                worksheet.cell(row=row_pointer, column=6, value=0)
                worksheet.cell(row=row_pointer, column=7, value=abs(difference))
                shortage_sum = difference * Decimal(product['price'])
                worksheet.cell(row=row_pointer, column=8, value=abs(shortage_sum))
                alls_shortage_count += abs(difference)
                alls_shortage_sum += abs(shortage_sum)

            row_pointer += 1

        # Итого
        worksheet.cell(row=row_pointer, column=4, value='Итого:')
        worksheet.cell(row=row_pointer, column=5, value=alls_surplus_count)
        worksheet.cell(row=row_pointer, column=6, value=alls_surplus_sum)
        worksheet.cell(row=row_pointer, column=7, value=alls_shortage_count)
        worksheet.cell(row=row_pointer, column=8, value=alls_shortage_sum)

    def process(self, *args, **kwargs):
        date = check_reports_directory()
        context = self._create_context()
        project_title = self.project.title.replace(' ', '_').replace('/', '_')
        filename = f'Отчет_по_расхождениям_{project_title}_{token_urlsafe(5)}.{"xlsx" if self.is_excel else "pdf"}'

        if self.is_excel:
            self.workbook = load_workbook(filename='api/v1/reports/excels/Differences.xlsx')
            self._generate_excel(context=context)
            self.workbook.save(f'{settings.MEDIA_ROOT}/reports/{date}/{filename}')
        else:
            res = render_to_string('reports/differences-report.html', context=context)
            output_dest = f'{settings.MEDIA_ROOT}/reports/{date}/{filename}'
            with open(output_dest, 'wb+') as ofile:
                pisa.CreatePDF(res, ofile)

        return f'{self.endpoint_pref}/media/reports/{date}/{filename}'
