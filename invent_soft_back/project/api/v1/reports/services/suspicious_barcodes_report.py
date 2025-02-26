from secrets import token_urlsafe

from django.conf import settings
from django.template.loader import render_to_string
from openpyxl import load_workbook
from xhtml2pdf import pisa

from api.v1.product.services import MergeProductTitleAndTitleAttrsService
from api.v1.reports.services.base_service import BaseReportService
from api.v1.reports.services.helpers.check_reports_directory import check_reports_directory
from apps import app
from apps.document.models import Document


@app.task
def generate_suspicious_barcodes_report_task(serializer_data: dict, endpoint_prefx: str):
    return SuspiciousBarcodesReportService(serializer_data, endpoint_prefx).process()


class SuspiciousBarcodesReportService(BaseReportService):  # noqa: WPS338
    """Сервис для генерации отчета Отчет по подозрительным ШК."""

    def __init__(self, serializer_data: dict, endpoint_pref: str):
        super().__init__(serializer_data, endpoint_pref)
        self.is_excel = serializer_data.get('excel')

    def _create_context(self):
        source_data = {}
        suspicious_barcodes_count = self.project.terminal_settings.suspicious_barcodes_amount

        for document in Document.objects.filter(zone__project=self.project, suspicious=True):
            for scan_product in document.counter_scan_task.scanned_products.all():
                if scan_product.product.title not in source_data:
                    source_data[scan_product.product.title] = {
                        'project_id': scan_product.product.project.id,
                        'product_id': scan_product.product.id,
                        'vendor_code': scan_product.product.vendor_code,
                        'barcode': scan_product.product.barcode,
                        'title': scan_product.product.title,
                        'count': scan_product.amount,
                        'zone': scan_product.task.zone.title,
                    }
                else:
                    source_data[scan_product.product.title]['count'] += scan_product.amount

        context = {'products': []}
        for prod_key in source_data.keys():
            if source_data[prod_key]['count'] >= suspicious_barcodes_count:  # noqa: WPS204
                context['products'].append({
                    'zone': source_data[prod_key]['zone'],
                    'vendor_code': source_data[prod_key]['vendor_code'],
                    'barcode': source_data[prod_key]['barcode'],
                    'title': MergeProductTitleAndTitleAttrsService(
                        project_id=source_data[prod_key]['project_id'],
                        product_id=source_data[prod_key]['product_id'],
                        product_title=source_data[prod_key]['title'],
                    ).process(),
                    'count': source_data[prod_key]['count'],
                })

        return context

    def _generate_excel(self, context):
        worksheet = self.workbook.worksheets[0]
        suspicious_barcodes_count = self.project.terminal_settings.suspicious_barcodes_amount
        row_pointer = 3
        for product in context['products']:
            if product['count'] >= suspicious_barcodes_count:
                worksheet.cell(row=row_pointer, column=1, value=product['zone'])
                worksheet.cell(row=row_pointer, column=2, value=product['vendor_code'])
                worksheet.cell(row=row_pointer, column=3, value=product['barcode'])
                worksheet.cell(row=row_pointer, column=4, value=product['title'])
                worksheet.cell(row=row_pointer, column=5, value=product['count'])
                row_pointer += 1

    def process(self, *args, **kwargs):
        date = check_reports_directory()
        context = self._create_context()
        project_title = self.project.title.replace(' ', '_').replace('/', '_')
        filename = f'Подозрительные_ШК_{project_title}_{token_urlsafe(5)}.{"xlsx" if self.is_excel else "pdf"}'

        if self.is_excel:
            self.workbook = load_workbook(filename='api/v1/reports/excels/SuspiciousBarcodes.xlsx')
            self._generate_excel(context=context)
            self.workbook.save(f'{settings.MEDIA_ROOT}/reports/{date}/{filename}')
        else:
            res = render_to_string('reports/suspicious-barcodes.html', context=context)
            output_dest = f'{settings.MEDIA_ROOT}/reports/{date}/{filename}'
            with open(output_dest, 'wb+') as ofile:
                pisa.CreatePDF(res, ofile)

        return f'{self.endpoint_pref}/media/reports/{date}/{filename}'
