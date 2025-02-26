from secrets import token_urlsafe

from django.conf import settings
from django.template.loader import render_to_string
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.pagebreak import Break
from xhtml2pdf import pisa

from api.v1.product.services import MergeProductTitleAndTitleAttrsService
from api.v1.reports.services.base_service import BaseReportService
from api.v1.reports.services.helpers.check_reports_directory import check_reports_directory
from api.v1.reports.services.helpers.number_to_text import NumberToTextService
from apps import app
from apps.document.models import Document, DocumentStatusChoices


@app.task
def generate_inv_three_report_task(serializer_data: dict, endpoint_prefx: str):
    return InvThreeReportService(serializer_data, endpoint_prefx).process()


class InvThreeReportService(BaseReportService):  # noqa: WPS338
    """Сервис для генерации отчета ИНВ-3."""

    def __init__(self, serializer_data: dict, endpoint_pref: str) -> None:
        super().__init__(serializer_data, endpoint_pref)
        self.group_by = serializer_data.get('group_by')
        self.include = serializer_data.get('include')
        self.is_excel = serializer_data.get('excel')
        self.workbook = None

    def _create_context(self):  # noqa: WPS231
        products = {}
        for document in Document.objects.filter(zone__project=self.project, status=DocumentStatusChoices.READY):
            for scanned_product in document.counter_scan_task.scanned_products.all():
                if (  # noqa: WPS337
                    self.include == 'all'  # noqa: WPS222
                    or self.include == 'only_identified' and scanned_product.product.title != 'Неизвестный товар'
                    or self.include == 'only_unidentified' and scanned_product.product.title == 'Неизвестный товар'
                    or self.include == 'found_by_product_code' and scanned_product.added_by_product_code is True
                ):
                    if scanned_product.product.barcode not in products:
                        code = scanned_product.product.barcode
                        if self.group_by == 'by_product_code':
                            code = scanned_product.product.vendor_code
                        products[scanned_product.product.barcode] = {
                            'code': code,
                            'title': MergeProductTitleAndTitleAttrsService(
                                project_id=scanned_product.product.project.id,
                                product_id=scanned_product.product.id,
                                product_title=scanned_product.product.title,
                            ).process(),
                            'price': float(scanned_product.product.price),
                            'amount': float(scanned_product.amount),
                        }
                    else:
                        products[scanned_product.product.barcode]['amount'] += float(scanned_product.amount)
        products = [item for item in products.values()]  # noqa: C416
        products = sorted(products, key=lambda x: x['code'])

        source_data = {
            'pages': [
                {
                    'products': [],
                    'page_serial_numbers': 0,
                    'page_number_units': 0,
                    'page_actual_amount': 0,
                    'number_page': 2,
                },
            ],
            'total_serial_numbers': 0,
            'total_number_units': 0,
            'total_actual_amount': 0,
            'total_number_page': 3,
        }
        products_on_page_counter = 0
        titles_on_page_length = 0
        page = 0
        number_page = 2
        for _, product in enumerate(products, start=1):  # noqa: WPS468
            if len(product.get('title')) > 29:
                product['title'] = product.get('title')[:30] + ' ' + product.get('title')[30:301]

            titles_on_page_length += len(product.get('title'))
            if not self.is_excel and (products_on_page_counter == 10 or titles_on_page_length > 750):
                source_data['pages'].append({
                    'products': [],
                    'page_serial_numbers': 0,
                    'page_number_units': 0,
                    'page_actual_amount': 0,
                    'number_page': 2,
                })
                products_on_page_counter = 0
                titles_on_page_length = 0
                page += 1
                number_page += 1
                source_data['pages'][page]['number_page'] = number_page
                source_data['total_number_page'] += 1
            elif self.is_excel and (products_on_page_counter == 10 or titles_on_page_length > 750):
                source_data['pages'].append({
                    'products': [],
                    'page_serial_numbers': 0,
                    'page_number_units': 0,
                    'page_actual_amount': 0,
                    'number_page': 2,
                })
                products_on_page_counter = 0
                titles_on_page_length = 0
                page += 1
                number_page += 1
                source_data['pages'][page]['number_page'] = number_page
                source_data['total_number_page'] += 1

            product['full_price'] = round(product.get('price') * product.get('amount'), 2)
            source_data['pages'][page]['products'].append(product)
            source_data['pages'][page]['page_serial_numbers'] += 1
            source_data['pages'][page]['page_number_units'] += product.get('amount')
            source_data['pages'][page]['page_actual_amount'] += (product.get('amount') * product.get('price'))
            source_data['total_serial_numbers'] += 1
            source_data['total_number_units'] += product.get('amount')
            source_data['total_actual_amount'] += (product.get('amount') * product.get('price'))
            products_on_page_counter += 1
        # Дополнительно пройтись в цикле, привести цифры в нормальный вид
        for page in source_data.get('pages'):
            for product in page.get('products'):
                if product.get('amount') % 1 == 0:  # noqa: WPS345
                    product['amount'] = int(product.get('amount'))
                else:
                    product['amount'] = round(product['amount'], 3)
            page['page_number_units'] = round(page['page_number_units'], 2)
            page['page_actual_amount'] = round(page['page_actual_amount'], 2)
            page['total_kop'] = int(round(page['page_actual_amount'] % 1, 2) * 100)  # noqa: WPS345

        source_data['total_actual_amount'] = round(source_data['total_actual_amount'], 3)
        source_data['total_actual_kop'] = int(round(source_data['total_actual_amount'] % 1, 2) * 100)  # noqa: WPS345

        # Перевести числа в слова
        source_data['total_serial_numbers_words'] = NumberToTextService().num2text(source_data['total_serial_numbers'])
        if not self.is_excel:
            source_data['total_serial_numbers_words'] += ''.join(
                ['_' for _ in range(105 - len(source_data['total_serial_numbers_words']))],
            )

        source_data['total_number_units_words'] = NumberToTextService().num2text(source_data['total_number_units'])
        if not self.is_excel:
            source_data['total_number_units_words'] += ''.join(
                ['_' for _ in range(102 - len(source_data['total_number_units_words']))],
            )

        rub, kop = NumberToTextService().decimal2text(
            value=source_data['total_actual_amount'], int_units=((u';', u';', u';'), 'm'),
        ).split(';')
        source_data['total_actual_amount_words_rub'] = rub
        source_data['total_actual_amount_words_kop'] = kop
        if not self.is_excel:
            source_data['total_actual_amount_words_rub'] += ''.join(['_' for _ in range(118 - len(rub))])

        for page in source_data['pages']:
            page['page_serial_numbers_words'] = NumberToTextService().num2text(page['page_serial_numbers'])
            if not self.is_excel:
                page['page_serial_numbers_words'] += ''.join(
                    ['_' for _ in range(118 - len(page['page_serial_numbers_words']))],
                )

            page['page_number_units_words'] = NumberToTextService().num2text(page['page_number_units'])
            if not self.is_excel:
                page['page_number_units_words'] += ''.join(
                    ['_' for _ in range(115 - len(page['page_number_units_words']))],
                )

            rub, kop = NumberToTextService().decimal2text(
                value=page['page_actual_amount'], int_units=((u';', u';', u';'), 'm'),
            ).split(';')
            page['page_actual_amount_words_rub'] = rub
            page['page_actual_amount_words_kop'] = kop
            if not self.is_excel:
                page['page_actual_amount_words_rub'] += ''.join(['_' for _ in range(130 - len(rub))])

        source_data['total_actual_amount'] = round(source_data['total_actual_amount'], 2)
        return source_data

    def _generate_excel(self, source_data):  # noqa: WPS213
        self.workbook = load_workbook(filename='api/v1/reports/excels/INV-3.xlsx')
        worksheet = self.workbook.worksheets[1]
        times_new_roman_font = Font(name='Times New Roman', size=7)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
        )
        top_border = Border(
            top=Side(style='thin', color='000000'),
        )
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        row_pointer = 1
        for page in source_data.get('pages'):
            worksheet.merge_cells(f'K{row_pointer}:M{row_pointer + 1}')
            worksheet.cell(row=row_pointer, column=11, value=f'{page.get("number_page")}-страница формы № ИНВ-3')
            worksheet.cell(row=row_pointer, column=11).alignment = Alignment(vertical='center')
            worksheet.cell(row=row_pointer, column=11).border = border
            worksheet.cell(row=row_pointer + 1, column=11).border = border
            worksheet.cell(row=row_pointer, column=12).border = border
            worksheet.cell(row=row_pointer, column=13).border = border
            worksheet.cell(row=row_pointer + 1, column=13).border = border
            worksheet.cell(row=row_pointer, column=11).font = times_new_roman_font
            row_pointer += 2
            worksheet.merge_cells(f'A{row_pointer}:A{row_pointer + 1}')
            worksheet.cell(row=row_pointer, column=1, value='№ п/п')
            worksheet.cell(row=row_pointer, column=1).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=1).border = border
            worksheet.cell(row=row_pointer, column=1).alignment = header_alignment
            worksheet.cell(row=row_pointer + 1, column=1).border = border

            worksheet.merge_cells(f'B{row_pointer}:B{row_pointer + 1}')
            worksheet.cell(row=row_pointer, column=2, value='Счет, субсчет')
            worksheet.cell(row=row_pointer, column=2).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=2).border = border
            worksheet.cell(row=row_pointer, column=2).alignment = header_alignment
            worksheet.cell(row=row_pointer + 1, column=2).border = border

            worksheet.merge_cells(f'C{row_pointer}:D{row_pointer}')
            worksheet.cell(row=row_pointer, column=3, value='Товарно-материальные ценности')
            worksheet.cell(row=row_pointer, column=3).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=3).border = border
            worksheet.cell(row=row_pointer, column=3).alignment = header_alignment
            worksheet.cell(row=row_pointer, column=4).border = border

            worksheet.cell(row=row_pointer + 1, column=3, value='Наименование, характеристика (вид, сорт, группа)')
            worksheet.cell(row=row_pointer + 1, column=3).font = times_new_roman_font
            worksheet.cell(row=row_pointer + 1, column=3).border = border
            worksheet.cell(row=row_pointer + 1, column=3).alignment = header_alignment

            worksheet.cell(row=row_pointer + 1, column=4, value='Код (номенклатурный номер)')
            worksheet.cell(row=row_pointer + 1, column=4).font = times_new_roman_font
            worksheet.cell(row=row_pointer + 1, column=4).border = border
            worksheet.cell(row=row_pointer + 1, column=4).alignment = header_alignment

            worksheet.merge_cells(f'E{row_pointer}:F{row_pointer}')
            worksheet.cell(row=row_pointer, column=5, value='Единица измерения')
            worksheet.cell(row=row_pointer, column=5).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=5).border = border
            worksheet.cell(row=row_pointer, column=5).alignment = header_alignment
            worksheet.cell(row=row_pointer, column=6).border = border

            worksheet.cell(row=row_pointer + 1, column=5, value='Код по ОКЕИ')
            worksheet.cell(row=row_pointer + 1, column=5).font = times_new_roman_font
            worksheet.cell(row=row_pointer + 1, column=5).border = border
            worksheet.cell(row=row_pointer + 1, column=5).alignment = header_alignment

            worksheet.cell(row=row_pointer + 1, column=6, value='Наименование')
            worksheet.cell(row=row_pointer + 1, column=6).font = times_new_roman_font
            worksheet.cell(row=row_pointer + 1, column=6).border = border
            worksheet.cell(row=row_pointer + 1, column=6).alignment = header_alignment

            worksheet.merge_cells(f'G{row_pointer}:G{row_pointer + 1}')
            worksheet.cell(row=row_pointer, column=7, value='Цена')
            worksheet.cell(row=row_pointer, column=7).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=7).border = border
            worksheet.cell(row=row_pointer, column=7).alignment = header_alignment

            worksheet.merge_cells(f'H{row_pointer}:I{row_pointer}')
            worksheet.cell(row=row_pointer, column=8, value='Номер')
            worksheet.cell(row=row_pointer, column=8).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=8).border = border
            worksheet.cell(row=row_pointer, column=8).alignment = header_alignment
            worksheet.cell(row=row_pointer, column=9).border = border
            worksheet.cell(row=row_pointer + 1, column=8, value='Инвентарный')
            worksheet.cell(row=row_pointer + 1, column=8).font = times_new_roman_font
            worksheet.cell(row=row_pointer + 1, column=8).border = border
            worksheet.cell(row=row_pointer + 1, column=8).alignment = header_alignment
            worksheet.cell(row=row_pointer + 1, column=9, value='Паспорта')
            worksheet.cell(row=row_pointer + 1, column=9).font = times_new_roman_font
            worksheet.cell(row=row_pointer + 1, column=9).border = border
            worksheet.cell(row=row_pointer + 1, column=9).alignment = header_alignment

            worksheet.merge_cells(f'J{row_pointer}:K{row_pointer}')
            worksheet.cell(row=row_pointer, column=10, value='Фактическое наличие')
            worksheet.cell(row=row_pointer, column=10).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=10).border = border
            worksheet.cell(row=row_pointer, column=10).alignment = header_alignment
            worksheet.cell(row=row_pointer, column=11).border = border
            worksheet.cell(row=row_pointer + 1, column=10, value='Количество')
            worksheet.cell(row=row_pointer + 1, column=10).font = times_new_roman_font
            worksheet.cell(row=row_pointer + 1, column=10).border = border
            worksheet.cell(row=row_pointer + 1, column=10).alignment = header_alignment
            worksheet.cell(row=row_pointer + 1, column=11, value='Сумма, руб. коп')
            worksheet.cell(row=row_pointer + 1, column=11).font = times_new_roman_font
            worksheet.cell(row=row_pointer + 1, column=11).border = border
            worksheet.cell(row=row_pointer + 1, column=11).alignment = header_alignment

            worksheet.merge_cells(f'L{row_pointer}:M{row_pointer}')
            worksheet.cell(row=row_pointer, column=12, value='По данным бух. Учета')
            worksheet.cell(row=row_pointer, column=12).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=12).border = border
            worksheet.cell(row=row_pointer, column=12).alignment = header_alignment
            worksheet.cell(row=row_pointer, column=13).border = border
            worksheet.cell(row=row_pointer + 1, column=12, value='Количество')
            worksheet.cell(row=row_pointer + 1, column=12).font = times_new_roman_font
            worksheet.cell(row=row_pointer + 1, column=12).border = border
            worksheet.cell(row=row_pointer + 1, column=12).alignment = header_alignment
            worksheet.cell(row=row_pointer + 1, column=13, value='Сумма, руб. коп')
            worksheet.cell(row=row_pointer + 1, column=13).font = times_new_roman_font
            worksheet.cell(row=row_pointer + 1, column=13).border = border
            worksheet.cell(row=row_pointer + 1, column=13).alignment = header_alignment

            for col in range(1, 14):
                worksheet.cell(row=row_pointer + 2, column=col).value = col
                worksheet.cell(row=row_pointer + 2, column=col).font = times_new_roman_font
                worksheet.cell(row=row_pointer + 2, column=col).border = border
                worksheet.cell(row=row_pointer + 2, column=col).alignment = header_alignment

            worksheet.row_dimensions[row_pointer].height = 45
            worksheet.row_dimensions[row_pointer + 1].height = 45
            row_pointer += 3

            for idx, product in enumerate(page.get('products'), start=1):
                worksheet.cell(row=row_pointer, column=1).font = times_new_roman_font
                worksheet.cell(row=row_pointer, column=1).border = border
                worksheet.cell(row=row_pointer, column=1, value=idx)
                worksheet.cell(row=row_pointer, column=1).alignment = header_alignment

                worksheet.cell(row=row_pointer, column=2).border = border
                worksheet.cell(row=row_pointer, column=2).alignment = header_alignment

                worksheet.cell(row=row_pointer, column=3).font = times_new_roman_font
                worksheet.cell(row=row_pointer, column=3).border = border
                worksheet.cell(row=row_pointer, column=3, value=product.get('title'))
                worksheet.cell(row=row_pointer, column=3).alignment = Alignment(wrap_text=True)

                worksheet.cell(row=row_pointer, column=4).font = times_new_roman_font
                worksheet.cell(row=row_pointer, column=4).border = border
                worksheet.cell(row=row_pointer, column=4, value=product.get('code'))

                worksheet.cell(row=row_pointer, column=5).border = border
                worksheet.cell(row=row_pointer, column=6).border = border

                worksheet.cell(row=row_pointer, column=7).font = times_new_roman_font
                worksheet.cell(row=row_pointer, column=7).border = border
                worksheet.cell(row=row_pointer, column=7, value=product.get('price'))
                worksheet.cell(row=row_pointer, column=7).alignment = header_alignment

                worksheet.cell(row=row_pointer, column=8).border = border
                worksheet.cell(row=row_pointer, column=9).border = border

                worksheet.cell(row=row_pointer, column=10).font = times_new_roman_font
                worksheet.cell(row=row_pointer, column=10).border = border
                worksheet.cell(row=row_pointer, column=10, value=product.get('amount'))
                worksheet.cell(row=row_pointer, column=10).alignment = header_alignment

                worksheet.cell(row=row_pointer, column=11).font = times_new_roman_font
                worksheet.cell(row=row_pointer, column=11).border = border
                worksheet.cell(row=row_pointer, column=11, value=product.get('full_price'))
                worksheet.cell(row=row_pointer, column=11).alignment = header_alignment

                worksheet.cell(row=row_pointer, column=12).border = border
                worksheet.cell(row=row_pointer, column=12).alignment = header_alignment
                worksheet.cell(row=row_pointer, column=13).border = border
                worksheet.cell(row=row_pointer, column=13).alignment = header_alignment

                row_pointer += 1

            worksheet.merge_cells(f'A{row_pointer}:I{row_pointer}')
            worksheet.cell(row=row_pointer, column=1, value='Итого:')
            worksheet.cell(row=row_pointer, column=1).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=1).border = border
            worksheet.cell(row=row_pointer, column=2).border = border
            worksheet.cell(row=row_pointer, column=3).border = border
            worksheet.cell(row=row_pointer, column=4).border = border
            worksheet.cell(row=row_pointer, column=5).border = border
            worksheet.cell(row=row_pointer, column=6).border = border
            worksheet.cell(row=row_pointer, column=7).border = border
            worksheet.cell(row=row_pointer, column=8).border = border
            worksheet.cell(row=row_pointer, column=9).border = border
            worksheet.cell(row=row_pointer, column=1).alignment = Alignment(horizontal='right')
            worksheet.cell(row=row_pointer, column=10, value=page.get('page_number_units'))
            worksheet.cell(row=row_pointer, column=10).border = border
            worksheet.cell(row=row_pointer, column=10).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=11, value=page.get('page_actual_amount'))
            worksheet.cell(row=row_pointer, column=11).border = border
            worksheet.cell(row=row_pointer, column=11).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=12).border = border
            worksheet.cell(row=row_pointer, column=12).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=13).border = border
            worksheet.cell(row=row_pointer, column=13).font = times_new_roman_font

            # Итого по странице
            row_pointer += 2
            worksheet.merge_cells(f'A{row_pointer}:M{row_pointer}')
            worksheet.cell(row=row_pointer, column=1, value='Итого по странице')
            worksheet.cell(row=row_pointer, column=1).font = times_new_roman_font
            row_pointer += 1

            # Кол-во порядковых номеров на странице
            worksheet.merge_cells(f'C{row_pointer}:M{row_pointer}')
            worksheet.merge_cells(f'D{row_pointer + 1}:M{row_pointer + 1}')
            value = 'а) количество порядковых номеров    ' + page.get('page_serial_numbers_words')
            worksheet.cell(row=row_pointer, column=3, value=value)
            worksheet.cell(row=row_pointer, column=3).font = times_new_roman_font
            for col in range(4, 14):
                worksheet.cell(row=row_pointer + 1, column=col).border = top_border
            worksheet.cell(row=row_pointer + 1, column=4, value='прописью').alignment = Alignment(horizontal='center')
            worksheet.cell(row=row_pointer + 1, column=4).font = times_new_roman_font
            row_pointer += 2

            # Общее кол-во единиц на странице
            worksheet.merge_cells(f'C{row_pointer}:M{row_pointer}')
            worksheet.merge_cells(f'D{row_pointer + 1}:M{row_pointer + 1}')
            value = 'б) общее количество единиц фактически    ' + page.get('page_number_units_words')
            worksheet.cell(row=row_pointer, column=3, value=value)
            worksheet.cell(row=row_pointer, column=3).font = times_new_roman_font
            worksheet.cell(row=row_pointer + 1, column=4, value='прописью').alignment = Alignment(horizontal='center')
            worksheet.cell(row=row_pointer + 1, column=4).font = times_new_roman_font
            for col in range(4, 14):
                worksheet.cell(row=row_pointer + 1, column=col).border = top_border
            row_pointer += 2

            # На сумму фактически
            worksheet.merge_cells(f'C{row_pointer}:J{row_pointer}')
            worksheet.merge_cells(f'D{row_pointer + 1}:M{row_pointer + 1}')
            value = 'в) на сумму фактически    ' + f'{page.get("page_actual_amount_words_rub")}'
            worksheet.cell(row=row_pointer, column=3, value=value)
            worksheet.cell(row=row_pointer, column=3).font = times_new_roman_font
            worksheet.cell(row=row_pointer + 1, column=4, value='прописью').alignment = Alignment(horizontal='center')
            worksheet.cell(row=row_pointer + 1, column=4).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=11, value='руб.').alignment = Alignment(horizontal='right')
            worksheet.cell(row=row_pointer, column=11).font = times_new_roman_font
            worksheet.cell(
                row=row_pointer,
                column=12,
                value=int(round(page.get('page_actual_amount') % 1, 2) * 100),  # noqa: WPS345
            ).alignment = Alignment(horizontal='center')
            worksheet.cell(row=row_pointer, column=12).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=13, value='коп.')
            worksheet.cell(row=row_pointer, column=13).font = times_new_roman_font
            for col in range(4, 14):
                worksheet.cell(row=row_pointer + 1, column=col).border = top_border

            row_pointer += 1
            next_page, _ = worksheet.page_breaks
            next_page.append(Break(row_pointer))
            row_pointer += 1

        worksheet = self.workbook.worksheets[2]
        worksheet.cell(row=1, column=29, value=f'{source_data.get("total_number_page")}-я страница формы № ИНВ-3')
        worksheet.cell(row=row_pointer, column=29).font = times_new_roman_font
        worksheet.cell(
            row=4,
            column=3,
            value=f'количество порядковых номеров:    {source_data.get("total_serial_numbers_words")}',
        )
        worksheet.cell(
            row=6,
            column=3,
            value=f'общее количество единиц фактически:    {source_data.get("total_number_units_words")}',
        )
        worksheet.cell(
            row=8,
            column=3,
            value=f'на сумму фактически:    {source_data.get("total_actual_amount_words_rub")}',
        )
        import math  # noqa: WPS433
        frac, whole = math.modf(source_data.get('total_actual_amount'))
        worksheet.cell(row=11, column=31, value=int(round(frac, 2) * 100)).alignment = Alignment(horizontal='center')

    def process(self, *args, **kwargs):
        date = check_reports_directory()
        context = self._create_context()
        project_title = self.project.title.replace(' ', '_').replace('/', '_')
        filename = f'ИНВ-3_{project_title}_{token_urlsafe(5)}.{"xlsx" if self.is_excel else "pdf"}'

        if self.is_excel:
            self._generate_excel(context)
            self.workbook.save(f'{settings.MEDIA_ROOT}/reports/{date}/{filename}')
        else:
            res = render_to_string('reports/inv-3.html', context=context)
            output_dest = f'{settings.MEDIA_ROOT}/reports/{date}/{filename}'
            with open(output_dest, 'wb+') as ofile:
                pisa.CreatePDF(res, ofile, encoding='utf-8')

        return f'{self.endpoint_pref}/media/reports/{date}/{filename}'
