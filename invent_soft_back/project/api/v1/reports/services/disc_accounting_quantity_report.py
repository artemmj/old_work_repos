import logging
from decimal import Decimal
from secrets import token_urlsafe

from django.conf import settings
from django.template.loader import render_to_string
from openpyxl import load_workbook
from xhtml2pdf import pisa

from api.v1.product.services import MergeProductTitleAndTitleAttrsService
from api.v1.reports.services.base_service import BaseReportService
from api.v1.reports.services.helpers.check_reports_directory import check_reports_directory
from apps import app
from apps.document.models import Document, DocumentStatusChoices
from apps.product.models import Product

logger = logging.getLogger('django')


@app.task
def generate_disc_accounting_quantity_report_task(serializer_data: dict, endpoint_prefx: str):
    return DiscAccountingQuantityReportService(serializer_data, endpoint_prefx).process()


class DiscAccountingQuantityReportService(BaseReportService):  # noqa: WPS338
    """Сервис для генерации отчета Расхождения с учетным кол-вом."""

    def __init__(self, serializer_data: dict, endpoint_pref: str):
        super().__init__(serializer_data, endpoint_pref)
        self.group_by = serializer_data.get('group_by')
        self.is_excel = serializer_data.get('excel')

    def _create_context(self):  # noqa: WPS231
        source_data = {}
        for product in Product.objects.filter(project=self.project, remainder__isnull=False):
            vendor_code = f'art_{product.barcode}'
            if product.vendor_code:
                vendor_code = product.vendor_code
            code = product.barcode if self.group_by == 'by_barcode' else vendor_code

            source_data[product] = {
                'code': code,
                'project_id': product.project.id,
                'product_id': product.id,
                'barcode': product.barcode,
                'vendor_code': vendor_code,
                'title': product.title,
                'price': float(product.price),
                'remainder': product.remainder,
                'zone': {},
                'fact': 0,
            }

        for document in Document.objects.filter(zone__project=self.project, status=DocumentStatusChoices.READY):
            for scan_prod in document.counter_scan_task.scanned_products.all():
                if scan_prod.product not in source_data:
                    source_data[scan_prod.product]['zone'] = {scan_prod.task.zone.serial_number: scan_prod.amount}
                    source_data[scan_prod.product]['fact'] = scan_prod.amount
                else:
                    source_data[scan_prod.product]['fact'] += scan_prod.amount
                    zone_ser_number = scan_prod.task.zone.serial_number
                    if zone_ser_number not in source_data[scan_prod.product]['zone']:
                        source_data[scan_prod.product]['zone'][zone_ser_number] = scan_prod.amount
                    else:
                        source_data[scan_prod.product]['zone'][zone_ser_number] += scan_prod.amount

        if self.group_by == 'by_barcode':
            source_data = dict(sorted(source_data.items(), key=lambda item: str(item[0].barcode)))
        elif self.group_by == 'by_product_code':
            source_data = dict(sorted(source_data.items(), key=lambda item: str(item[0].vendor_code)))

        context = {
            'products': [],
            'alls_fact': 0,
            'alls_uchet': 0,
            'alls_diff_amount': 0,
            'alls_price_amount': 0,
        }
        for product in source_data.keys():
            if source_data[product]['remainder'] == source_data[product]['fact']:
                continue

            if source_data[product]['remainder'] < 0:
                diff_amount = abs(source_data[product]['remainder'] - source_data[product]['fact'])
            else:
                diff_amount = source_data[product]['fact'] - source_data[product]['remainder']
            diff_price = diff_amount * Decimal(source_data[product]['price'])

            zones = ''
            for zone_num, amount in source_data[product]['zone'].items():  # noqa: WPS519
                zones += f'{zone_num}-{amount}, '

            vendor_code = source_data[product]['vendor_code']
            if len(vendor_code) > 9:
                vendor_code = vendor_code[:10] + ' ' + vendor_code[10:]

            context['products'].append({
                'barcode': source_data[product]['barcode'],
                'title': MergeProductTitleAndTitleAttrsService(
                    project_id=source_data[product]['project_id'],
                    product_id=source_data[product]['product_id'],
                    product_title=source_data[product]['title'],
                ).process(),
                'vendor_code': vendor_code,
                'zone': zones,
                'fact': round(source_data[product]['fact'], 3),
                'uchet': round(source_data[product]['remainder'], 3),
                'price': round(source_data[product]['price'], 2),
                'diff_amount': round(diff_amount, 3),
                'diff_price': round(diff_price, 2),
            })
            context['alls_fact'] += source_data[product]['fact']
            context['alls_uchet'] += source_data[product]['remainder']
            context['alls_diff_amount'] += diff_amount
            context['alls_price_amount'] += diff_price

        context['alls_fact'] = round(context['alls_fact'], 3)
        context['alls_uchet'] = round(context['alls_uchet'], 3)
        context['alls_diff_amount'] = round(context['alls_diff_amount'], 2)
        context['alls_price_amount'] = round(context['alls_price_amount'], 2)

        return context

    def _generate_excel(self, context):  # noqa: WPS213
        worksheet = self.workbook.worksheets[0]

        alls_fact = 0
        alls_uchet = 0
        alls_diff_amount = 0
        alls_price_amount = 0

        row_pointer = 2
        for idx, product in enumerate(context.get('products'), start=1):
            worksheet.cell(row=row_pointer, column=1, value=idx)
            worksheet.cell(row=row_pointer, column=2, value=product['barcode'])
            worksheet.cell(row=row_pointer, column=3, value=product['title'])
            worksheet.cell(row=row_pointer, column=4, value=product['vendor_code'])
            worksheet.cell(row=row_pointer, column=5, value=product['zone'])
            worksheet.cell(row=row_pointer, column=6, value=product['fact'])
            worksheet.cell(row=row_pointer, column=7, value=product['uchet'])
            worksheet.cell(row=row_pointer, column=8, value=product['price'])

            diff_amount = product['fact'] - product['uchet']
            worksheet.cell(row=row_pointer, column=9, value=diff_amount)
            diff_price = diff_amount * Decimal(product['price'])
            worksheet.cell(row=row_pointer, column=10, value=diff_price)

            alls_fact += product['fact']
            alls_uchet += product['uchet']
            alls_diff_amount += diff_amount
            alls_price_amount += diff_price

            row_pointer += 1

        # Итого
        worksheet.cell(row=row_pointer, column=5, value='Итого: ')
        worksheet.cell(row=row_pointer, column=6, value=alls_fact)
        worksheet.cell(row=row_pointer, column=7, value=alls_uchet)
        worksheet.cell(row=row_pointer, column=9, value=alls_diff_amount)
        worksheet.cell(row=row_pointer, column=10, value=alls_price_amount)

    def process(self, *args, **kwargs):
        date = check_reports_directory()
        context = self._create_context()
        filename = f'disc-accounting-quantity-report_{token_urlsafe(5)}.{"xlsx" if self.is_excel else "pdf"}'

        if self.is_excel:
            self.workbook = load_workbook(filename='api/v1/reports/excels/DiffRests.xlsx')
            self._generate_excel(context=context)
            self.workbook.save(f'{settings.MEDIA_ROOT}/reports/{date}/{filename}')
        else:
            res = render_to_string('reports/discrep_accounting_quantity.html', context=context)
            output_dest = f'{settings.MEDIA_ROOT}/reports/{date}/{filename}'
            with open(output_dest, 'wb+') as ofile:
                pisa.CreatePDF(res, ofile)

        return f'{self.endpoint_pref}/media/reports/{date}/{filename}'
