import operator
from datetime import datetime
from secrets import token_urlsafe

from django.conf import settings
from django.db.models import Sum
from django.template.loader import render_to_string
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from xhtml2pdf import pisa

from api.v1.reports.services.helpers.check_reports_directory import check_reports_directory
from apps import app
from apps.document.models import Document, DocumentStatusChoices
from apps.employee.models import Employee
from apps.helpers.services import AbstractService
from apps.task.models import TaskTypeChoices


@app.task
def generate_auditor_report_task(serializer_data: dict, endpoint_prefx: str):
    return AuditorReportService(serializer_data, endpoint_prefx).process()


class AuditorReportService(AbstractService):  # noqa: WPS214
    def __init__(self, serializer_data: dict, endpoint_pref: str):
        self.endpoint_pref = endpoint_pref
        self.employee = Employee.objects.get(pk=serializer_data.get('auditor'))
        self.is_excel = serializer_data.get('excel')
        self.task_type = serializer_data.get('task_type')

    def process(self, *args, **kwargs):  # noqa: WPS231
        date = check_reports_directory()
        employee_name = self.employee.username.replace('/', '_').replace(' ', '_')
        filename = ''

        if self.task_type == TaskTypeChoices.AUDITOR:
            filename = f'Отчет_аудитора_{employee_name}_{token_urlsafe(5)}.{"xlsx" if self.is_excel else "pdf"}'
            context = self._prepare_content_for_auditor_report()
            if self.is_excel:
                self.workbook = load_workbook(filename='api/v1/reports/excels/Auditor.xlsx')
                self.worksheet = self.workbook.worksheets[0]
                self._create_xlsx_with_auditor_report(context=context)
                self.workbook.save(f'{settings.MEDIA_ROOT}/reports/{date}/{filename}')
            else:
                res = render_to_string('reports/auditor-report.html', context=context)
                output_dest = f'{settings.MEDIA_ROOT}/reports/{date}/{filename}'
                with open(output_dest, 'wb+') as ofile:
                    pisa.CreatePDF(res, ofile)

        elif self.task_type == TaskTypeChoices.AUDITOR_CONTROLLER:
            filename = f'Отчет_аудитора_ук_{employee_name}_{token_urlsafe(5)}.{"xlsx" if self.is_excel else "pdf"}'
            context = self._prepare_content_for_auditor_controller_report()
            if self.is_excel:
                self.workbook = load_workbook(filename='api/v1/reports/excels/AuditorController.xlsx')
                self.worksheet = self.workbook.worksheets[0]
                self._create_xlsx_with_auditor_controller_report(context=context)
                self.workbook.save(f'{settings.MEDIA_ROOT}/reports/{date}/{filename}')
            else:
                res = render_to_string('reports/auditor-controller-report.html', context=context)
                output_dest = f'{settings.MEDIA_ROOT}/reports/{date}/{filename}'
                with open(output_dest, 'wb+') as ofile:
                    pisa.CreatePDF(res, ofile)

        return f'{self.endpoint_pref}/media/reports/{date}/{filename}'

    def _prepare_content_for_auditor_controller_report(self):  # noqa: WPS118
        content_for_report = {
            'object': f'{self.employee.project.title}, {self.employee.project.address}',
            'date': datetime.now().strftime('%d-%m-%Y'),
            'auditor': self.employee,
            'zones': [],
            'statistic': {
                'verified_zones_count': 0,
                'verified_zones_percent': 0,
                'rescanned_zones_count': 0,
            },
        }

        documents = Document.objects.filter(
            auditor_controller_task__employee=self.employee,
            zone__project=self.employee.project,
            status=DocumentStatusChoices.READY,
        )
        for document in documents:
            counter_scan_task_result = int(document.counter_scan_task.result) if document.counter_scan_task else ' '
            controller_tasks_result = int(document.controller_task.result) if document.controller_task else '-'
            auditor_controller_tasks_result = (
                int(document.auditor_controller_task.result)
                if document.auditor_controller_task
                else ' '
            )
            is_result_discrepancy = (
                '-'
                if controller_tasks_result == auditor_controller_tasks_result
                else 'Да'
            )

            content_for_report['zones'].append({
                'zone_number': document.zone.serial_number,
                'counter_scan_task_result': counter_scan_task_result,
                'controller_tasks_result': controller_tasks_result,
                'auditor_controller_tasks_result': auditor_controller_tasks_result,
                'is_result_discrepancy': is_result_discrepancy,
            })

        content_for_report['statistic']['verified_zones_count'] = len(content_for_report['zones'])
        content_for_report['statistic']['percentage_of_scanned_areas'] = self._calculate_percentage_of_scanned_areas(
            employee_documents=documents,
        )

        content_for_report['zones'].sort(key=operator.itemgetter('zone_number'))
        return content_for_report

    def _prepare_content_for_auditor_report(self):
        content_for_report = {
            'object': f'{self.employee.project.title}, {self.employee.project.address}',
            'date': datetime.now().strftime('%d-%m-%Y'),
            'auditor': self.employee,
            'zones': [],
            'statistic': {
                'verified_zones_count': 0,
                'verified_zones_percent': 0,
                'rescanned_zones_count': 0,
            },
        }

        documents = Document.objects.filter(
            auditor_task__employee=self.employee,
            zone__project=self.employee.project,
            status=DocumentStatusChoices.READY,
        )

        for document in documents:
            auditor_task_result = int(document.auditor_task.result) if document.auditor_task else ' '
            counter_scan_task_result = int(document.counter_scan_task.result) if document.counter_scan_task else ' '
            is_replace_specification = 'Да' if document.is_replace_specification else '-'
            is_difference_between_scanned_products = 'Да' if self._compare_scanned_products_for(document) else '-'

            content_for_report['zones'].append({
                'zone_number': document.zone.serial_number,
                'auditor_task_result': auditor_task_result,
                'counter_scan_task_result': counter_scan_task_result,
                'quantity_discrepancy': auditor_task_result - counter_scan_task_result,
                'is_replace_specification': is_replace_specification,
                'is_difference_between_scanned_products': is_difference_between_scanned_products,
            })

        content_for_report['statistic']['verified_zones_count'] = len(content_for_report['zones'])
        content_for_report['statistic'][
            'docs_with_specification_change_count'
        ] = self._calculate_docs_with_specification_change_count_for(documents)
        content_for_report['statistic']['scanned_products_count'] = self._scanned_products_count_for(documents)
        content_for_report['statistic']['percentage_of_scanned_areas'] = self._calculate_percentage_of_scanned_areas(
            documents,
        )

        content_for_report['zones'].sort(key=operator.itemgetter('zone_number'))
        return content_for_report

    def _create_xlsx_with_auditor_report(self, context):  # noqa: WPS213
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        times_new_roman_font = Font(name='Times New Roman', size=14)

        self.worksheet.cell(row=3, column=1, value=f'Объект: {context["object"]}')
        self.worksheet.cell(row=3, column=1).font = times_new_roman_font
        self.worksheet.cell(row=3, column=1).border = thin_border

        self.worksheet.cell(row=5, column=1, value=f'Дата: {context["date"]}')
        self.worksheet.cell(row=5, column=1).font = times_new_roman_font
        self.worksheet.cell(row=5, column=1).border = thin_border

        self.worksheet.cell(row=7, column=1, value=f'Аудитор: {context["auditor"]} (Внутренний аудитор)')
        self.worksheet.cell(row=7, column=1).font = times_new_roman_font
        self.worksheet.cell(row=7, column=1).border = thin_border

        row_counter = 11
        for idx, zone in enumerate(context['zones'], start=1):
            self.worksheet.cell(row=row_counter, column=1, value=idx)
            self.worksheet.cell(row=row_counter, column=1).font = times_new_roman_font
            self.worksheet.cell(row=row_counter, column=1).border = thin_border
            self.worksheet.cell(row=row_counter, column=1).alignment = Alignment(horizontal='center')

            self.worksheet.cell(row=row_counter, column=2, value=zone['zone_number'])
            self.worksheet.cell(row=row_counter, column=2).font = times_new_roman_font
            self.worksheet.cell(row=row_counter, column=2).border = thin_border
            self.worksheet.cell(row=row_counter, column=2).alignment = Alignment(horizontal='center')

            self.worksheet.cell(row=row_counter, column=3, value=zone['auditor_task_result'])
            self.worksheet.cell(row=row_counter, column=3).font = times_new_roman_font
            self.worksheet.cell(row=row_counter, column=3).border = thin_border
            self.worksheet.cell(row=row_counter, column=3).alignment = Alignment(horizontal='center')

            self.worksheet.cell(row=row_counter, column=4, value=zone['counter_scan_task_result'])
            self.worksheet.cell(row=row_counter, column=4).font = times_new_roman_font
            self.worksheet.cell(row=row_counter, column=4).border = thin_border
            self.worksheet.cell(row=row_counter, column=4).alignment = Alignment(horizontal='center')

            self.worksheet.cell(row=row_counter, column=5, value=zone['quantity_discrepancy'])
            self.worksheet.cell(row=row_counter, column=5).font = times_new_roman_font
            self.worksheet.cell(row=row_counter, column=5).border = thin_border
            self.worksheet.cell(row=row_counter, column=5).alignment = Alignment(horizontal='center')

            self.worksheet.cell(row=row_counter, column=6, value=zone['is_difference_between_scanned_products'])
            self.worksheet.cell(row=row_counter, column=6).font = times_new_roman_font
            self.worksheet.cell(row=row_counter, column=6).border = thin_border
            self.worksheet.cell(row=row_counter, column=6).alignment = Alignment(horizontal='center')

            self.worksheet.cell(row=row_counter, column=7, value=zone['is_replace_specification'])
            self.worksheet.cell(row=row_counter, column=7).font = times_new_roman_font
            self.worksheet.cell(row=row_counter, column=7).border = thin_border
            self.worksheet.cell(row=row_counter, column=7).alignment = Alignment(horizontal='center')

            row_counter += 1

        row_counter += 1
        self.worksheet.merge_cells(f'A{row_counter}:D{row_counter}')
        self.worksheet.cell(row=row_counter, column=1, value='Статистика аудитора:')
        self.worksheet.cell(row=row_counter, column=1).font = times_new_roman_font
        self.worksheet.cell(row=row_counter, column=1).border = thin_border
        self.worksheet.cell(row=row_counter, column=2).border = thin_border
        self.worksheet.cell(row=row_counter, column=3).border = thin_border
        self.worksheet.cell(row=row_counter, column=4).border = thin_border
        self.worksheet.cell(row=row_counter, column=1).alignment = Alignment(horizontal='center')
        row_counter += 1

        self.worksheet.merge_cells(f'A{row_counter}:C{row_counter}')
        self.worksheet.cell(row=row_counter, column=1, value='Кол-во зон, проверенных аудитором:')
        self.worksheet.cell(row=row_counter, column=1).font = times_new_roman_font
        self.worksheet.cell(row=row_counter, column=1).border = thin_border
        self.worksheet.cell(row=row_counter, column=2).border = thin_border
        self.worksheet.cell(row=row_counter, column=3).border = thin_border
        self.worksheet.cell(row=row_counter, column=1).alignment = Alignment(horizontal='center')
        self.worksheet.cell(row=row_counter, column=4, value=context['statistic']['verified_zones_count'])
        self.worksheet.cell(row=row_counter, column=4).font = times_new_roman_font
        self.worksheet.cell(row=row_counter, column=4).border = thin_border
        self.worksheet.cell(row=row_counter, column=4).alignment = Alignment(horizontal='center')
        row_counter += 1

        self.worksheet.merge_cells(f'A{row_counter}:C{row_counter}')
        self.worksheet.cell(row=row_counter, column=1, value='Общее кол-во отсканированных позиций:')
        self.worksheet.cell(row=row_counter, column=1).font = times_new_roman_font
        self.worksheet.cell(row=row_counter, column=1).border = thin_border
        self.worksheet.cell(row=row_counter, column=2).border = thin_border
        self.worksheet.cell(row=row_counter, column=3).border = thin_border
        self.worksheet.cell(row=row_counter, column=1).alignment = Alignment(horizontal='center')
        self.worksheet.cell(row=row_counter, column=4, value=context['statistic']['scanned_products_count'])
        self.worksheet.cell(row=row_counter, column=4).font = times_new_roman_font
        self.worksheet.cell(row=row_counter, column=4).border = thin_border
        self.worksheet.cell(row=row_counter, column=4).alignment = Alignment(horizontal='center')
        row_counter += 1

        self.worksheet.merge_cells(f'A{row_counter}:C{row_counter}')
        self.worksheet.cell(row=row_counter, column=1, value='% отсканированных зон:')
        self.worksheet.cell(row=row_counter, column=1).font = times_new_roman_font
        self.worksheet.cell(row=row_counter, column=1).border = thin_border
        self.worksheet.cell(row=row_counter, column=2).border = thin_border
        self.worksheet.cell(row=row_counter, column=3).border = thin_border
        self.worksheet.cell(row=row_counter, column=1).alignment = Alignment(horizontal='center')
        self.worksheet.cell(
            row=row_counter, column=4, value=f'{context["statistic"]["percentage_of_scanned_areas"]} %',
        )
        self.worksheet.cell(row=row_counter, column=4).font = times_new_roman_font
        self.worksheet.cell(row=row_counter, column=4).border = thin_border
        self.worksheet.cell(row=row_counter, column=4).alignment = Alignment(horizontal='center')
        row_counter += 1

        self.worksheet.merge_cells(f'A{row_counter}:C{row_counter}')
        self.worksheet.cell(row=row_counter, column=1, value='Кол-во док-тов с заменой спецификации:')
        self.worksheet.cell(row=row_counter, column=1).font = times_new_roman_font
        self.worksheet.cell(row=row_counter, column=1).border = thin_border
        self.worksheet.cell(row=row_counter, column=2).border = thin_border
        self.worksheet.cell(row=row_counter, column=3).border = thin_border
        self.worksheet.cell(row=row_counter, column=1).alignment = Alignment(horizontal='center')
        self.worksheet.cell(
            row=row_counter, column=4, value=context['statistic']['docs_with_specification_change_count'],
        )
        self.worksheet.cell(row=row_counter, column=4).font = times_new_roman_font
        self.worksheet.cell(row=row_counter, column=4).border = thin_border
        self.worksheet.cell(row=row_counter, column=4).alignment = Alignment(horizontal='center')

    def _create_xlsx_with_auditor_controller_report(self, context):  # noqa: WPS213
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        times_new_roman_font = Font(name='Times New Roman', size=14)

        self.worksheet.cell(row=3, column=1, value=f'Объект: {context["object"]}')
        self.worksheet.cell(row=3, column=1).font = times_new_roman_font
        self.worksheet.cell(row=3, column=1).border = thin_border

        self.worksheet.cell(row=5, column=1, value=f'Дата: {context["date"]}')
        self.worksheet.cell(row=5, column=1).font = times_new_roman_font
        self.worksheet.cell(row=5, column=1).border = thin_border

        self.worksheet.cell(row=7, column=1, value=f'Аудитор: {context["auditor"]} (Внутренний аудитор УК)')
        self.worksheet.cell(row=7, column=1).font = times_new_roman_font
        self.worksheet.cell(row=7, column=1).border = thin_border

        row_counter = 11
        for idx, zone in enumerate(context['zones'], start=1):
            self.worksheet.cell(row=row_counter, column=1, value=idx)
            self.worksheet.cell(row=row_counter, column=1).font = times_new_roman_font
            self.worksheet.cell(row=row_counter, column=1).border = thin_border
            self.worksheet.cell(row=row_counter, column=1).alignment = Alignment(horizontal='center')

            self.worksheet.cell(row=row_counter, column=2, value=zone['zone_number'])
            self.worksheet.cell(row=row_counter, column=2).font = times_new_roman_font
            self.worksheet.cell(row=row_counter, column=2).border = thin_border
            self.worksheet.cell(row=row_counter, column=2).alignment = Alignment(horizontal='center')

            self.worksheet.cell(row=row_counter, column=3, value=zone['counter_scan_task_result'])
            self.worksheet.cell(row=row_counter, column=3).font = times_new_roman_font
            self.worksheet.cell(row=row_counter, column=3).border = thin_border
            self.worksheet.cell(row=row_counter, column=3).alignment = Alignment(horizontal='center')

            self.worksheet.cell(row=row_counter, column=4, value=zone['controller_tasks_result'])
            self.worksheet.cell(row=row_counter, column=4).font = times_new_roman_font
            self.worksheet.cell(row=row_counter, column=4).border = thin_border
            self.worksheet.cell(row=row_counter, column=4).alignment = Alignment(horizontal='center')

            self.worksheet.cell(row=row_counter, column=5, value=zone['auditor_controller_tasks_result'])
            self.worksheet.cell(row=row_counter, column=5).font = times_new_roman_font
            self.worksheet.cell(row=row_counter, column=5).border = thin_border
            self.worksheet.cell(row=row_counter, column=5).alignment = Alignment(horizontal='center')

            self.worksheet.cell(row=row_counter, column=6, value=zone['is_result_discrepancy'])
            self.worksheet.cell(row=row_counter, column=6).font = times_new_roman_font
            self.worksheet.cell(row=row_counter, column=6).border = thin_border
            self.worksheet.cell(row=row_counter, column=6).alignment = Alignment(horizontal='center')

            row_counter += 1

        row_counter += 1
        self.worksheet.merge_cells(f'A{row_counter}:D{row_counter}')
        self.worksheet.cell(row=row_counter, column=1, value='Статистика аудитора:')
        self.worksheet.cell(row=row_counter, column=1).font = times_new_roman_font
        self.worksheet.cell(row=row_counter, column=1).border = thin_border
        self.worksheet.cell(row=row_counter, column=2).border = thin_border
        self.worksheet.cell(row=row_counter, column=3).border = thin_border
        self.worksheet.cell(row=row_counter, column=4).border = thin_border
        self.worksheet.cell(row=row_counter, column=1).alignment = Alignment(horizontal='center')
        row_counter += 1

        self.worksheet.merge_cells(f'A{row_counter}:C{row_counter}')
        self.worksheet.cell(row=row_counter, column=1, value='Кол-во зон, проверенных аудитором:')
        self.worksheet.cell(row=row_counter, column=1).font = times_new_roman_font
        self.worksheet.cell(row=row_counter, column=1).border = thin_border
        self.worksheet.cell(row=row_counter, column=2).border = thin_border
        self.worksheet.cell(row=row_counter, column=3).border = thin_border
        self.worksheet.cell(row=row_counter, column=1).alignment = Alignment(horizontal='center')
        self.worksheet.cell(row=row_counter, column=4, value=context['statistic']['verified_zones_count'])
        self.worksheet.cell(row=row_counter, column=4).font = times_new_roman_font
        self.worksheet.cell(row=row_counter, column=4).border = thin_border
        self.worksheet.cell(row=row_counter, column=4).alignment = Alignment(horizontal='center')
        row_counter += 1

        self.worksheet.merge_cells(f'A{row_counter}:C{row_counter}')
        self.worksheet.cell(row=row_counter, column=1, value='% отсканированных зон:')
        self.worksheet.cell(row=row_counter, column=1).font = times_new_roman_font
        self.worksheet.cell(row=row_counter, column=1).border = thin_border
        self.worksheet.cell(row=row_counter, column=2).border = thin_border
        self.worksheet.cell(row=row_counter, column=3).border = thin_border
        self.worksheet.cell(row=row_counter, column=1).alignment = Alignment(horizontal='center')
        self.worksheet.cell(
            row=row_counter, column=4, value=f'{context["statistic"]["percentage_of_scanned_areas"]} %',
        )
        self.worksheet.cell(row=row_counter, column=4).font = times_new_roman_font
        self.worksheet.cell(row=row_counter, column=4).border = thin_border
        self.worksheet.cell(row=row_counter, column=4).alignment = Alignment(horizontal='center')
        row_counter += 1

        self.worksheet.merge_cells(f'A{row_counter}:C{row_counter}')
        self.worksheet.cell(row=row_counter, column=1, value='Кол-во пересканированных зон:')
        self.worksheet.cell(row=row_counter, column=1).font = times_new_roman_font
        self.worksheet.cell(row=row_counter, column=1).border = thin_border
        self.worksheet.cell(row=row_counter, column=2).border = thin_border
        self.worksheet.cell(row=row_counter, column=3).border = thin_border
        self.worksheet.cell(row=row_counter, column=1).alignment = Alignment(horizontal='center')
        self.worksheet.cell(row=row_counter, column=4, value=0)
        self.worksheet.cell(row=row_counter, column=4).font = times_new_roman_font
        self.worksheet.cell(row=row_counter, column=4).border = thin_border
        self.worksheet.cell(row=row_counter, column=4).alignment = Alignment(horizontal='center')

    def _compare_scanned_products_for(self, document: Document) -> bool:
        counter_scanned_products = document.counter_scan_task.scanned_products.all()
        auditor_scanned_products = document.auditor_task.scanned_products.all()

        different_quantity_items = []
        for counter_scanned_product in counter_scanned_products:
            auditor_scanned_product = (
                auditor_scanned_products
                .filter(product__id=counter_scanned_product.product.id)
                .first()
            )

            if not auditor_scanned_product:
                continue

            if (  # noqa: WPS337
                counter_scanned_product.amount != auditor_scanned_product.amount
                and counter_scanned_product.product.id == auditor_scanned_product.product.id
            ):
                different_quantity_items.append(counter_scanned_product)

        return True if different_quantity_items else False  # noqa: WPS502

    def _calculate_docs_with_specification_change_count_for(self, documents) -> int:  # noqa: WPS118
        return documents.filter(is_replace_specification=True).count()

    def _scanned_products_count_for(self, documents) -> int:
        total_scanned_products_amount = documents.aggregate(
            Sum('auditor_task__scanned_products__amount'),
        )['auditor_task__scanned_products__amount__sum']
        if total_scanned_products_amount is None:
            return 0
        else:
            return int(total_scanned_products_amount) if documents else 0  # noqa:  WPS503

    def _calculate_percentage_of_scanned_areas(self, employee_documents):
        all_documents_in_ready_status = Document.objects.filter(
            zone__project=self.employee.project,
            status=DocumentStatusChoices.READY,
        )
        if all_documents_in_ready_status.count() == 0:
            return 0
        else:
            return round(employee_documents.count() / all_documents_in_ready_status.count() * 100, 2)  # noqa:  WPS503
