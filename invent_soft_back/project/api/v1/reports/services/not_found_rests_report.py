from secrets import token_urlsafe

from django.conf import settings
from django.template.loader import render_to_string
from openpyxl import load_workbook
from xhtml2pdf import pisa

from api.v1.reports.services.base_service import BaseReportService
from api.v1.reports.services.helpers.check_reports_directory import check_reports_directory
from apps import app
from apps.product.models import ScannedProduct


@app.task
def generate_not_found_rests_report_task(serializer_data: dict, endpoint_prefx: str):
    return NotFoundRestsReportService(serializer_data, endpoint_prefx).process()


class NotFoundRestsReportService(BaseReportService):  # noqa: WPS338
    """Сервис для генерации отчета Ненайденные остатки."""

    def __init__(self, serializer_data: dict, endpoint_pref: str) -> None:
        super().__init__(serializer_data, endpoint_pref)
        self.group_by = serializer_data['group_by']
        self.discrepancy_in = serializer_data['discrepancy_in']
        self.more_than = serializer_data['more_than']
        self.is_excel = serializer_data['excel']

    def _create_context_for_report(self):  # noqa: WPS231
        source_data = {}

        if self.group_by == 'by_product_code':
            products = ScannedProduct.objects.filter(product__project=self.project).order_by('product__vendor_code')
        elif self.group_by == 'by_barcode':
            products = ScannedProduct.objects.filter(product__project=self.project).order_by('product__barcode')

        for product in products:
            if product.product in source_data:
                source_data[product.product]['fact'] += product.amount
            else:
                source_data[product.product] = {
                    'barcode': product.product.barcode,
                    'title': product.product.title,
                    'price': float(product.product.price),
                    'uchet': product.product.remainder,
                    'fact': product.amount,
                }

        context = {
            'products': [],
            'alls_fact_sum': 0,
            'alls_uchet_sum': 0,
            'alls_diff_count_sum': 0,
            'alls_diff_price_sum': 0,
        }

        for prodkey in source_data.keys():
            count_diff = source_data[prodkey]['uchet'] - source_data[prodkey]['fact']  # noqa: WPS204
            product_data = {}
            if source_data[prodkey]['fact'] == 0 and count_diff > self.more_than:
                product_data = source_data[prodkey]
                product_data['diff_count'] = source_data[prodkey]['uchet']
                product_data['diff_price'] = source_data[prodkey]['price'] * source_data[prodkey]['uchet']

                context['products'].append(product_data)

                context['alls_fact_sum'] += product_data['fact']
                context['alls_uchet_sum'] += product_data['uchet']
                context['alls_diff_count_sum'] += product_data['uchet']
                context['alls_diff_price_sum'] += product_data['price'] * product_data['uchet']

        if self.group_by == 'by_barcode':
            context['gr'] = 'По штрихкоду'
        elif self.group_by == 'by_product_code':
            context['gr'] = 'По коду товара'

        return context

    def generate_excel(self, context):  # noqa: WPS213
        self.workbook = load_workbook(filename='api/v1/reports/excels/NotFoundRests.xlsx')
        worksheet = self.workbook.worksheets[0]
        worksheet.cell(row=2, column=1, value=f'Группировка: {context["gr"]}')

        alls_fact_sum = 0
        alls_uchet_sum = 0
        alls_diff_count_sum = 0
        alls_diff_price_sum = 0

        row_pointer = 5
        for idx, product in enumerate(context.get('products'), start=1):
            worksheet.cell(row=row_pointer, column=1, value=idx)
            worksheet.cell(row=row_pointer, column=2, value=product['barcode'])
            worksheet.cell(row=row_pointer, column=3, value=product['title'])
            worksheet.cell(row=row_pointer, column=6, value=product['fact'])
            worksheet.cell(row=row_pointer, column=7, value=product['uchet'])  # noqa: WPS204
            worksheet.cell(row=row_pointer, column=8, value=product['price'])
            worksheet.cell(row=row_pointer, column=9, value=product['uchet'])
            worksheet.cell(row=row_pointer, column=10, value=product['price'] * product['uchet'])

            alls_fact_sum += product['fact']
            alls_uchet_sum += product['uchet']
            alls_diff_count_sum += product['uchet']
            alls_diff_price_sum += product['price'] * product['uchet']

            row_pointer += 1

        worksheet.cell(row=row_pointer, column=5, value='Итого: ')
        worksheet.cell(row=row_pointer, column=6, value=alls_fact_sum)
        worksheet.cell(row=row_pointer, column=7, value=alls_uchet_sum)
        worksheet.cell(row=row_pointer, column=9, value=alls_diff_count_sum)
        worksheet.cell(row=row_pointer, column=10, value=alls_diff_price_sum)

    def process(self, *args, **kwargs):
        date = check_reports_directory()
        context = self._create_context_for_report()
        project_title = self.project.title.replace('/', '_').replace(' ', '_')
        filename = f'Ненайденные_остатки_{project_title}_{token_urlsafe(5)}.{"xlsx" if self.is_excel else "pdf"}'

        if self.is_excel:
            self.generate_excel(context=context)
            self.workbook.save(f'{settings.MEDIA_ROOT}/reports/{date}/{filename}')
        else:
            res = render_to_string('reports/not-found-rests.html', context=context)
            output_dest = f'{settings.MEDIA_ROOT}/reports/{date}/{filename}'
            with open(output_dest, 'wb+') as ofile:
                pisa.CreatePDF(res, ofile)

        return f'{self.endpoint_pref}/media/reports/{date}/{filename}'
