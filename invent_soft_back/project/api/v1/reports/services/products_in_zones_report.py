from secrets import token_urlsafe

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

from api.v1.reports.services.base_service import BaseReportService
from api.v1.reports.services.helpers.check_reports_directory import check_reports_directory
from apps import app
from apps.product.models import ScannedProduct


@app.task
def generate_products_in_zones_report(serializer_data: dict, endpoint_prefix: str):
    return ProductsInZonesReportService(serializer_data, endpoint_prefix).process()


class ProductsInZonesReportService(BaseReportService):
    """Сервис для генерации отчета Товары в зонах."""

    def __init__(self, serializer_data: dict, endpoint_pref: str) -> None:
        super().__init__(serializer_data, endpoint_pref)
        self.workbook = load_workbook(filename='api/v1/reports/excels/ProductsInZones.xlsx')

    def process(self, *args, **kwargs):
        worksheet = self.workbook.worksheets[0]

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        scanned_products = ScannedProduct.objects.filter(task__zone__project=self.project)
        row_pointer = 3
        for sc_product in scanned_products:
            worksheet.cell(row=row_pointer, column=1, value=sc_product.product.vendor_code)
            worksheet.cell(row=row_pointer, column=2, value=sc_product.task.zone.title)
            worksheet.cell(row=row_pointer, column=3, value=sc_product.amount)
            worksheet.cell(row=row_pointer, column=4, value=sc_product.product.barcode)
            worksheet.cell(row=row_pointer, column=5, value=sc_product.product.title)

            for i in range(1, 6):
                worksheet.cell(row=row_pointer, column=i).border = thin_border

            row_pointer += 1

        date = check_reports_directory()
        project_title = self.project.title.replace('/', '_').replace(' ', '_')
        filename = f'Товары_в_зонах_{project_title}_{token_urlsafe(5)}.xlsx'
        self.workbook.save(f'{settings.MEDIA_ROOT}/reports/{date}/{filename}')
        return f'{self.endpoint_pref}/media/reports/{date}/{filename}'
