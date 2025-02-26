from decimal import Decimal
from secrets import token_urlsafe

from django.conf import settings
from django.template.loader import render_to_string
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from xhtml2pdf import pisa

from api.v1.product.services import MergeProductTitleAndTitleAttrsService
from api.v1.reports.services.base_service import BaseReportService
from api.v1.reports.services.helpers.check_reports_directory import check_reports_directory
from apps import app
from apps.document.models import Document, DocumentStatusChoices
from apps.product.models import Product


@app.task
def generate_discrepancies_report_task(serializer_data: dict, endpoint_prefx: str):
    return DiscrepanciesReportService(serializer_data, endpoint_prefx).process()


class DiscrepanciesReportService(BaseReportService):  # noqa: WPS338
    """Сервис для генерации отчета L&G сличительная (Ведомость расхождений)."""

    def __init__(self, serializer_data: dict, endpoint_pref: str):
        super().__init__(serializer_data, endpoint_pref)
        self.is_excel = serializer_data.get('excel')

    def _create_context_for_report(self, *args, **kwargs):  # noqa: WPS231
        source_data = {}
        for product in Product.objects.filter(project=self.project, remainder__isnull=False):
            vendor_code = f'art_{product.barcode}'
            if product.vendor_code:
                vendor_code = product.vendor_code

            source_data[product] = {
                'barcode': product.barcode,
                'title': MergeProductTitleAndTitleAttrsService(
                    project_id=product.project.id,
                    product_id=product.id,
                    product_title=product.title,
                ).process(),
                'vendor_code': vendor_code,
                'price': float(product.price),
                'remainder': product.remainder,
                'zone': {},
                'fact': 0,
            }

        for document in Document.objects.filter(zone__project=self.project, status=DocumentStatusChoices.READY):
            for scan_prod in document.counter_scan_task.scanned_products.all():
                if scan_prod.product not in source_data:
                    source_data[scan_prod.product]['zone'] = {scan_prod.task.zone.serial_number: scan_prod.amount}
                    source_data[scan_prod.product]['fact'] = scan_prod.amount
                else:
                    source_data[scan_prod.product]['fact'] += scan_prod.amount
                    zone_ser_number = scan_prod.task.zone.serial_number
                    if zone_ser_number not in source_data[scan_prod.product]['zone']:
                        source_data[scan_prod.product]['zone'][zone_ser_number] = scan_prod.amount
                    else:
                        source_data[scan_prod.product]['zone'][zone_ser_number] += scan_prod.amount

        context = {
            'products': [],
            'alls_diff_amount': 0,
            'alls_diff_price': 0,
        }

        for prodkey in source_data.keys():
            if source_data[prodkey]['remainder'] == source_data[prodkey]['fact']:
                continue

            if source_data[prodkey]['remainder'] < 0:
                diff_amount = abs(source_data[prodkey]['remainder'] - source_data[prodkey]['fact'])
            else:
                diff_amount = source_data[prodkey]['fact'] - source_data[prodkey]['remainder']
            diff_price = diff_amount * Decimal(source_data[prodkey]['price'])

            zones = ''
            for zone_num, amount in source_data[prodkey]['zone'].items():  # noqa: WPS519
                zones += f'{zone_num}-{round(amount, 3)}, '

            vendor_code = source_data[prodkey]['vendor_code']
            if len(vendor_code) > 9:
                vendor_code = vendor_code[:9] + ' ' + vendor_code[9:]

            context['products'].append({
                'vendor_code': vendor_code,
                'title': source_data[prodkey]['title'],
                'barcode': source_data[prodkey]['barcode'],
                'price': round(source_data[prodkey]['price'], 2),
                'zone': zones,
                'uchet': round(source_data[prodkey]['remainder'], 3),
                'fact': round(source_data[prodkey]['fact'], 3),
                'diff_amount': round(diff_amount, 3),
                'diff_price': round(diff_price, 2),
            })
            context['alls_diff_amount'] += abs(diff_amount)
            context['alls_diff_price'] += diff_price

        context['alls_diff_amount'] = round(context['alls_diff_amount'], 3)
        context['alls_diff_price'] = round(context['alls_diff_price'], 2)
        return context

    def _generate_excel(self, context):  # noqa: WPS213
        worksheet = self.workbook.worksheets[0]

        alls_diff_amount = 0
        alls_diff_price = 0

        times_new_roman_font = Font(name='Times New Roman', size=9)
        alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        row_pointer = 4
        for idx, product in enumerate(context.get('products'), start=1):
            worksheet.cell(row=row_pointer, column=1, value=idx)
            worksheet.cell(row=row_pointer, column=1, value=idx).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=1, value=idx).alignment = alignment

            worksheet.cell(row=row_pointer, column=2, value=product['vendor_code'])
            worksheet.cell(row=row_pointer, column=2).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=2).alignment = alignment

            worksheet.cell(row=row_pointer, column=4, value=product['title'])
            worksheet.cell(row=row_pointer, column=4).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=4).alignment = alignment

            worksheet.cell(row=row_pointer, column=6, value=product['barcode'])
            worksheet.cell(row=row_pointer, column=6).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=6).alignment = alignment

            worksheet.cell(row=row_pointer, column=9, value=product['price'])
            worksheet.cell(row=row_pointer, column=9).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=9).alignment = alignment

            worksheet.cell(row=row_pointer, column=10, value=product['uchet'])
            worksheet.cell(row=row_pointer, column=10).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=10).alignment = alignment

            worksheet.cell(row=row_pointer, column=11, value=product['fact'])
            worksheet.cell(row=row_pointer, column=11).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=11).alignment = alignment

            worksheet.cell(row=row_pointer, column=14, value=product['zone'])
            worksheet.cell(row=row_pointer, column=14).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=14).alignment = alignment

            diff_amount = product['uchet'] - product['fact']
            worksheet.cell(row=row_pointer, column=12, value=abs(diff_amount))
            worksheet.cell(row=row_pointer, column=12).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=12).alignment = alignment

            diff_price = abs(diff_amount) * Decimal(product['price'])
            worksheet.cell(row=row_pointer, column=13, value=diff_price)
            worksheet.cell(row=row_pointer, column=13).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=13).alignment = alignment

            alls_diff_amount += abs(diff_amount)
            alls_diff_price += diff_price

            row_pointer += 1

        worksheet.cell(row=row_pointer, column=11, value='Итого: ')
        worksheet.cell(row=row_pointer, column=11).font = times_new_roman_font
        worksheet.cell(row=row_pointer, column=11).alignment = alignment

        worksheet.cell(row=row_pointer, column=12, value=alls_diff_amount)
        worksheet.cell(row=row_pointer, column=12).font = times_new_roman_font
        worksheet.cell(row=row_pointer, column=12).alignment = alignment

        worksheet.cell(row=row_pointer, column=13, value=alls_diff_price)
        worksheet.cell(row=row_pointer, column=13).font = times_new_roman_font
        worksheet.cell(row=row_pointer, column=13).alignment = alignment

    def process(self, *args, **kwargs):
        date = check_reports_directory()
        context = self._create_context_for_report()
        project_title = self.project.title.replace(' ', '_').replace('/', '_')
        filename = f'Ведомость_расхождений_{project_title}_{token_urlsafe(5)}.{"xlsx" if self.is_excel else "pdf"}'

        if self.is_excel:
            self.workbook = load_workbook(filename='api/v1/reports/excels/DiscrepanciesList.xlsx')
            self._generate_excel(context=context)
            self.workbook.save(f'{settings.MEDIA_ROOT}/reports/{date}/{filename}')
        else:
            res = render_to_string('reports/discrepancies-report.html', context=context)
            output_dest = f'{settings.MEDIA_ROOT}/reports/{date}/{filename}'
            with open(output_dest, 'wb+') as ofile:
                pisa.CreatePDF(res, ofile)

        return f'{self.endpoint_pref}/media/reports/{date}/{filename}'
