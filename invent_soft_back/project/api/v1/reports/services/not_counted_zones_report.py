from datetime import datetime
from secrets import token_urlsafe

from django.conf import settings
from django.template.loader import render_to_string
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from xhtml2pdf import pisa

from api.v1.reports.services.base_service import BaseReportService
from api.v1.reports.services.helpers.check_reports_directory import check_reports_directory
from apps import app
from apps.zone.models import ZoneStatusChoices


@app.task
def generate_not_counted_zones_report_task(serializer_data: dict, endpoint_prefx: str):
    return NotCountedZonesReportService(serializer_data, endpoint_prefx).process()


class NotCountedZonesReportService(BaseReportService):  # noqa: WPS338
    """Сервис для генерации отчета Непросчитанные зоны."""

    def __init__(self, serializer_data: dict, endpoint_pref: str):
        super().__init__(serializer_data, endpoint_pref)
        self.is_excel = serializer_data.get('excel')

    def _create_context(self):
        context = {'zones': {}}
        for zone in self.project.zones.filter(status=ZoneStatusChoices.NOT_READY).order_by('code'):
            st_name = zone.storage_name if zone.storage_name else 'Без названия склада'
            if st_name not in context['zones']:
                context['zones'][st_name] = {
                    'count': 1,
                    'zones': [zone.title],
                }
            else:
                context['zones'][st_name]['count'] += 1
                context['zones'][st_name]['zones'].append(zone.title)

        return context

    def _generate_excel(self, context):
        worksheet = self.workbook.worksheets[0]
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        times_new_roman_font = Font(name='Times New Roman', size=14)

        date = datetime.now().strftime('%Y-%m-%d %H:%M')
        context = {'date_title': f'{date} -- {self.project.title}', 'storages': {}}
        for zone in self.project.zones.filter(status=ZoneStatusChoices.NOT_READY).order_by('code'):
            st_name = zone.storage_name if zone.storage_name else 'Без названия склада'
            if st_name not in context['storages']:  # noqa: WPS204
                context['storages'][st_name] = {
                    'count': 1,
                    'zones': [zone.title],
                }
            else:
                context['storages'][st_name]['count'] += 1
                context['storages'][st_name]['zones'].append(zone.title)

        worksheet.cell(row=1, column=1, value=context['date_title'])

        row_pointer = 6
        col_pointer = 1
        idx = 0

        for storage_name in context['storages'].keys():

            worksheet.merge_cells(f'A{row_pointer}:C{row_pointer}')
            worksheet.cell(row=row_pointer, column=col_pointer, value=f'Наименование склада: {storage_name}')
            worksheet.cell(  # noqa: WPS204
                row=row_pointer,
                column=col_pointer,
            ).alignment = Alignment(horizontal='center')
            worksheet.cell(row=row_pointer, column=col_pointer).border = thin_border
            worksheet.cell(row=row_pointer, column=col_pointer).font = times_new_roman_font

            worksheet.cell(row=row_pointer, column=col_pointer + 1).border = thin_border
            worksheet.cell(row=row_pointer, column=col_pointer + 2).border = thin_border
            worksheet.cell(row=row_pointer, column=col_pointer + 3).border = thin_border

            worksheet.merge_cells(f'E{row_pointer}:G{row_pointer}')
            worksheet.cell(
                row=row_pointer,
                column=5,
                value=f'Количество: {context["storages"][storage_name]["count"]}',
            )
            worksheet.cell(row=row_pointer, column=5).alignment = Alignment(horizontal='center')
            worksheet.cell(row=row_pointer, column=5).font = times_new_roman_font

            worksheet.cell(row=row_pointer, column=5).border = thin_border
            worksheet.cell(row=row_pointer, column=6).border = thin_border
            worksheet.cell(row=row_pointer, column=7).border = thin_border

            row_pointer += 2

            for zone in context['storages'][storage_name]['zones']:
                worksheet.cell(row=row_pointer, column=col_pointer, value=zone)
                worksheet.cell(row=row_pointer, column=col_pointer).font = times_new_roman_font
                worksheet.cell(row=row_pointer, column=col_pointer).border = thin_border
                worksheet.cell(row=row_pointer, column=col_pointer).alignment = Alignment(horizontal='center')
                col_pointer += 2
                idx += 1

                if idx == 4:
                    row_pointer += 2
                    col_pointer = 1
                    idx = 0

            idx = 0
            row_pointer += 3
            col_pointer = 1

    def process(self, *args, **kwargs):
        date = check_reports_directory()
        context = self._create_context()
        project_title = self.project.title.replace(' ', '_').replace('/', '_')
        filename = f'Непросчитанные_зоны_{project_title}_{token_urlsafe(5)}.{"xlsx" if self.is_excel else "pdf"}'

        if self.is_excel:
            self.workbook = load_workbook(filename='api/v1/reports/excels/NotCountedZones.xlsx')
            self._generate_excel(context=context)
            self.workbook.save(f'{settings.MEDIA_ROOT}/reports/{date}/{filename}')
        else:
            res = render_to_string('reports/not-counted-zones.html', context=context)
            output_dest = f'{settings.MEDIA_ROOT}/reports/{date}/{filename}'
            with open(output_dest, 'wb+') as ofile:
                pisa.CreatePDF(res, ofile)

        return f'{self.endpoint_pref}/media/reports/{date}/{filename}'
