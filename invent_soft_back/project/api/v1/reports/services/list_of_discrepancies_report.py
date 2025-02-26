import logging
from decimal import Decimal
from secrets import token_urlsafe

from django.conf import settings
from django.template.loader import render_to_string
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from xhtml2pdf import pisa

from api.v1.product.services import MergeProductTitleAndTitleAttrsService
from api.v1.reports.services.base_service import BaseReportService
from api.v1.reports.services.helpers.check_reports_directory import check_reports_directory
from apps import app
from apps.document.models import Document, DocumentStatusChoices
from apps.product.models import Product

logger = logging.getLogger('django')


@app.task
def generate_list_of_discrepancies_report_task(serializer_data: dict, endpoint_prefx: str):
    return ListOfDiscrepanciesReportService(serializer_data, endpoint_prefx).process()


class ListOfDiscrepanciesReportService(BaseReportService):  # noqa: WPS338
    """Сервис для генерации отчета Ведомость расхождения Концепт Груп."""

    def __init__(self, serializer_data: dict, endpoint_pref: str):
        super().__init__(serializer_data, endpoint_pref)
        self.only_discrepancies = serializer_data.get('only_discrepancies')
        self.is_excel = serializer_data.get('excel')

    def _create_context(self, *args, **kwargs):  # noqa: WPS231
        source_data = {}
        for product in Product.objects.filter(project=self.project, remainder__isnull=False):
            source_data[product] = {
                'project_id': product.project.id,
                'product_id': product.id,
                'barcode': product.barcode,
                'title': product.title,
                'vendor_code': product.vendor_code,
                'price': float(product.price),
                'remainder': product.remainder,
                'zone': [],
                'fact': 0,
            }

        for document in Document.objects.filter(zone__project=self.project, status=DocumentStatusChoices.READY):
            for scan_product in document.counter_scan_task.scanned_products.all():
                if scan_product.product not in source_data:
                    source_data[scan_product.product] = {
                        'barcode': scan_product.product.barcode,
                        'vendor_code': scan_product.product.vendor_code,
                        'title': scan_product.product.title,
                        'price': float(scan_product.product.price),
                        'zone': [scan_product.task.zone.serial_number],
                        'uchet': scan_product.product.remainder,
                        'fact': scan_product.amount,
                    }
                    source_data[scan_product.product]['zone'] = [scan_product.task.zone.serial_number]
                    source_data[scan_product.product]['fact'] = scan_product.amount
                else:
                    source_data[scan_product.product]['fact'] += scan_product.amount
                    zone_ser_number = scan_product.task.zone.serial_number
                    if zone_ser_number not in source_data[scan_product.product]['zone']:
                        source_data[scan_product.product]['zone'].append(zone_ser_number)

        context = {
            'project': self.project.title,
            'products': [],
            'alls_fact_count': 0,
            'alls_diff_count': 0,
            'alls_diff_price': 0,
        }

        for product in source_data.keys():
            if source_data[product]['remainder'] < 0:
                diff_amount = abs(source_data[product]['remainder'] - source_data[product]['fact'])
            else:
                diff_amount = source_data[product]['fact'] - source_data[product]['remainder']
            diff_price = diff_amount * Decimal(source_data[product]['price'])

            zones = ', '.join([str(zone_num) for zone_num in source_data[product]['zone']])

            if diff_amount == 0:
                continue

            context['products'].append({
                'barcode': source_data[product]['barcode'],
                'vendor_code': source_data[product]['vendor_code'],
                'title': MergeProductTitleAndTitleAttrsService(
                    project_id=source_data[product]['project_id'],
                    product_id=source_data[product]['product_id'],
                    product_title=source_data[product]['title'],
                ).process(),
                'price': round(source_data[product]['price'], 2),
                'zone': zones,
                'uchet': round(source_data[product]['remainder'], 3),
                'fact': round(source_data[product]['fact'], 3),
                'diff_count': round(source_data[product]['fact'] - source_data[product]['remainder'], 3),
                'diff_price': round(diff_amount * Decimal(source_data[product]['price']), 2),
            })

            context['alls_fact_count'] += source_data[product]['fact']
            context['alls_diff_count'] += diff_amount
            context['alls_diff_price'] += diff_price

        context['alls_fact_count'] = round(context['alls_fact_count'], 3)
        context['alls_diff_count'] = round(context['alls_diff_count'], 3)
        context['alls_diff_price'] = round(context['alls_diff_price'], 2)

        return context

    def _generate_excel(self, context):  # noqa: WPS213
        worksheet = self.workbook.worksheets[0]

        times_new_roman_font = Font(name='Times New Roman', size=9)
        alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        row_pointer = 4
        for idx, product in enumerate(context.get('products'), start=1):
            worksheet.cell(row=row_pointer, column=1, value=idx)
            worksheet.cell(row=row_pointer, column=1).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=1).alignment = alignment

            worksheet.cell(row=row_pointer, column=2, value=product['barcode'])
            worksheet.cell(row=row_pointer, column=2).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=2).alignment = alignment

            worksheet.cell(row=row_pointer, column=3, value=product['vendor_code'])
            worksheet.cell(row=row_pointer, column=3).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=3).alignment = alignment

            worksheet.cell(row=row_pointer, column=5, value=product['title'])
            worksheet.cell(row=row_pointer, column=5).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=5).alignment = alignment

            worksheet.cell(row=row_pointer, column=6, value=product['price'])
            worksheet.cell(row=row_pointer, column=6).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=6).alignment = alignment

            worksheet.cell(row=row_pointer, column=9, value=product['uchet'])
            worksheet.cell(row=row_pointer, column=9).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=9).alignment = alignment

            worksheet.cell(row=row_pointer, column=10, value=product['fact'])
            worksheet.cell(row=row_pointer, column=10).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=10).alignment = alignment

            worksheet.cell(row=row_pointer, column=13, value=product['zone'])
            worksheet.cell(row=row_pointer, column=13).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=13).alignment = alignment

            diff_count = product['fact'] - product['uchet']
            worksheet.cell(row=row_pointer, column=11, value=diff_count)
            worksheet.cell(row=row_pointer, column=11).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=11).alignment = alignment

            diff_price = diff_count * Decimal(product['price'])
            worksheet.cell(row=row_pointer, column=12, value=diff_price)
            worksheet.cell(row=row_pointer, column=12).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=12).alignment = alignment

            row_pointer += 1

        worksheet.merge_cells(f'A{row_pointer}:E{row_pointer}')
        worksheet.cell(row=row_pointer, column=1, value='Итого:')
        worksheet.cell(row=row_pointer, column=1).font = times_new_roman_font
        worksheet.cell(row=row_pointer, column=1).alignment = Alignment(horizontal='right', vertical='center')

        worksheet.cell(row=row_pointer, column=10, value=context['alls_fact_count'])
        worksheet.cell(row=row_pointer, column=10).font = times_new_roman_font
        worksheet.cell(row=row_pointer, column=10).alignment = alignment

        worksheet.cell(row=row_pointer, column=11, value=context['alls_diff_count'])
        worksheet.cell(row=row_pointer, column=11).font = times_new_roman_font
        worksheet.cell(row=row_pointer, column=11).alignment = alignment

        worksheet.cell(row=row_pointer, column=12, value=context['alls_diff_price'])
        worksheet.cell(row=row_pointer, column=12).font = times_new_roman_font
        worksheet.cell(row=row_pointer, column=12).alignment = alignment

    def process(self, *args, **kwargs):
        date = check_reports_directory()
        context = self._create_context()
        project_title = self.project.title.replace('/', '_').replace(' ', '_')
        filename = f'Ведомость_расхождений_{project_title}_{token_urlsafe(5)}.{"xlsx" if self.is_excel else "pdf"}'

        if self.is_excel:
            self.workbook = load_workbook(filename='api/v1/reports/excels/ListOfDiscrepancies.xlsx')
            self._generate_excel(context=context)
            self.workbook.save(f'{settings.MEDIA_ROOT}/reports/{date}/{filename}')
        else:
            res = render_to_string('reports/list-of-discrepancies.html', context=context)
            output_dest = f'{settings.MEDIA_ROOT}/reports/{date}/{filename}'
            with open(output_dest, 'wb+') as ofile:
                pisa.CreatePDF(res, ofile)

        return f'{self.endpoint_pref}/media/reports/{date}/{filename}'
