import logging
from secrets import token_urlsafe

from django.conf import settings
from django.template.loader import render_to_string
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from xhtml2pdf import pisa

from api.v1.reports.services.base_service import BaseReportService
from api.v1.reports.services.helpers.check_reports_directory import check_reports_directory
from apps import app
from apps.document.models import Document, DocumentStatusChoices

logger = logging.getLogger('django')


@app.task
def generate_collation_statement_report_task(serializer_data: dict, endpoint_prefx: str):
    return CollationStatementTMCReportService(serializer_data, endpoint_prefx).process()


class CollationStatementTMCReportService(BaseReportService):  # noqa: WPS338
    """Сервис для генерации отчета Счислительная ведомость ТМЦ по зонам."""

    def __init__(self, serializer_data: dict, endpoint_pref: str) -> None:
        super().__init__(serializer_data, endpoint_pref)
        self.group_by = serializer_data['group_by']
        self.discrepancy_in = serializer_data['discrepancy_in']
        self.more_than = serializer_data['more_than']
        self.is_excel = serializer_data.get('excel')

    def _create_context(self):  # noqa: WPS231
        discrepancy_in_price = True if self.discrepancy_in == 'price' else False  # noqa: WPS502

        prds_src = {}
        for document in Document.objects.filter(status=DocumentStatusChoices.READY):
            for scan_product in document.counter_scan_task.scanned_products.all():
                # В зависимости от параметра включать в отчет либо все товары, либо только те, которые с ценой
                if discrepancy_in_price and not scan_product.product.price:
                    continue
                if scan_product.product.title not in prds_src:
                    product = scan_product.product
                    vendor_code = product.vendor_code if product.vendor_code else f'art_{product.barcode}'
                    prds_src[scan_product.product.title] = {
                        'barcode': product.barcode,
                        'vendor_code': vendor_code,
                        'title': product.title,
                        'price': product.price,
                        'zones': {},
                        'fact': scan_product.amount,
                        'uchet': product.remainder,
                    }
                else:
                    prds_src[scan_product.product.title]['fact'] += scan_product.amount

                if scan_product.task.zone.title not in prds_src[scan_product.product.title]['zones']:
                    prds_src[scan_product.product.title]['zones'][scan_product.task.zone.title] = scan_product.amount
                else:
                    prds_src[scan_product.product.title]['zones'][scan_product.task.zone.title] += scan_product.amount

        alls_fact_sum = 0
        alls_uchet_sum = 0
        alls_pm_sum = 0
        alls_pm_price_sum = 0
        alls_pm_price_abs_sum = 0

        context = {'products': []}
        for prod_key in prds_src.keys():
            prd = prds_src[prod_key].copy()
            difference = prd['fact'] - prd['uchet']
            if difference != 0 and abs(difference) > self.more_than:
                prd['pm'] = prd['fact'] - prd['uchet']
                prd['pm_price'] = prd['pm'] * prd['price']
                prd['pm_price_abs'] = abs(prd['pm'] * prd['price'])

                alls_fact_sum += prd['fact']
                alls_uchet_sum += prd['uchet']
                alls_pm_sum += prd['pm']
                alls_pm_price_sum += prd['pm_price']
                alls_pm_price_abs_sum += prd['pm_price_abs']

            context['products'].append(prd)

        context['alls_fact_sum'] = alls_fact_sum
        context['alls_uchet_sum'] = alls_uchet_sum
        context['alls_pm_sum'] = alls_pm_sum
        context['alls_pm_price_sum'] = alls_pm_price_sum
        context['alls_pm_price_abs_sum'] = alls_pm_price_abs_sum
        context['products'] = sorted(context['products'], key=lambda x: x['vendor_code'])
        if self.group_by == 'by_barcode':
            context['products'] = sorted(context['products'], key=lambda x: x['barcode'])

        return context

    def _generate_excel(self, context):  # noqa: WPS213
        self.workbook = load_workbook(filename='api/v1/reports/excels/DiffZone.xlsx')
        worksheet = self.workbook.worksheets[0]

        row_pointer = 4
        for prd in context['products']:
            worksheet.cell(row=row_pointer, column=1, value=prd['barcode'])
            worksheet.cell(row=row_pointer, column=2, value=prd['vendor_code'])
            worksheet.cell(row=row_pointer, column=3, value=prd['title'])
            worksheet.cell(row=row_pointer, column=4, value=prd['fact'])
            worksheet.cell(row=row_pointer, column=5, value=prd['uchet'])
            worksheet.cell(row=row_pointer, column=6, value=prd['price'])
            worksheet.cell(row=row_pointer, column=7, value=prd['pm'])
            worksheet.cell(row=row_pointer, column=8, value=prd['pm_price'])
            worksheet.cell(row=row_pointer, column=9, value=abs(prd['pm_price_abs']))
            for i in range(1, 10):
                worksheet.cell(row=row_pointer, column=i).alignment = Alignment(horizontal='center')

            row_pointer += 1

            for zone, count in prd['zones'].items():
                worksheet.merge_cells(f'A{row_pointer}:C{row_pointer}')
                worksheet.cell(row=row_pointer, column=1, value=zone)
                worksheet.cell(row=row_pointer, column=1).alignment = Alignment(horizontal='right')
                worksheet.cell(row=row_pointer, column=4, value=count)
                worksheet.cell(row=row_pointer, column=4).alignment = Alignment(horizontal='left')
                row_pointer += 1

        worksheet.merge_cells(f'A{row_pointer}:C{row_pointer}')
        worksheet.cell(row=row_pointer, column=1, value='Итого: ')
        worksheet.cell(row=row_pointer, column=1).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_pointer, column=4, value=context['alls_fact_sum'])
        worksheet.cell(row=row_pointer, column=4).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_pointer, column=5, value=context['alls_uchet_sum'])
        worksheet.cell(row=row_pointer, column=5).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_pointer, column=7, value=context['alls_pm_sum'])
        worksheet.cell(row=row_pointer, column=7).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_pointer, column=8, value=context['alls_pm_price_sum'])
        worksheet.cell(row=row_pointer, column=8).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_pointer, column=9, value=context['alls_pm_price_abs_sum'])
        worksheet.cell(row=row_pointer, column=9).alignment = Alignment(horizontal='center')

    def process(self, *args, **kwargs):
        date = check_reports_directory()
        context = self._create_context()
        project_title = self.project.title.replace(' ', '_').replace('/', '_')
        filename = (
            f'Счислительная_ведомость_ТМЦ_по_зонам_'  # noqa: F541
            f'{project_title}_{token_urlsafe(5)}.{"xlsx" if self.is_excel else "pdf"}'  # noqa: WPS326
        )

        if self.is_excel:
            self._generate_excel(context=context)
            self.workbook.save(f'{settings.MEDIA_ROOT}/reports/{date}/{filename}')
        else:
            res = render_to_string('reports/collation-statement-tmc.html', context=context)
            output_dest = f'{settings.MEDIA_ROOT}/reports/{date}/{filename}'
            with open(output_dest, 'wb+') as ofile:
                pisa.CreatePDF(res, ofile)

        return f'{self.endpoint_pref}/media/reports/{date}/{filename}'
