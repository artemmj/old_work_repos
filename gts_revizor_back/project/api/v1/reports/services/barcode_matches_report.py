import logging
from secrets import token_urlsafe

from django.conf import settings
from django.template.loader import render_to_string
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from xhtml2pdf import pisa

from api.v1.product.services import MergeProductTitleAndTitleAttrsService
from api.v1.reports.services.base_service import BaseReportService
from api.v1.reports.services.helpers.check_reports_directory import check_reports_directory
from api.v1.reports.services.helpers.escape_xlsx_chars import escape_xlsx_char
from apps import app
from apps.document.models import DocumentStatusChoices
from apps.task.models import TaskStatusChoices, TaskTypeChoices
from apps.zone.models import ZoneStatusChoices
from apps.zone.services.get_zones import GetZonesService

logger = logging.getLogger('django')


@app.task
def generate_barcode_matches_report_task(serializer_data: dict, endpoint_prefx: str, by_barcode: bool):
    return BarcodeMatchesReportService(serializer_data, endpoint_prefx, by_barcode).process()


class BarcodeMatchesReportService(BaseReportService):  # noqa: WPS338
    """Сервис для генерации отчета Совпадение ШК по зонам."""

    def __init__(self, serializer_data: dict, endpoint_pref: str, by_barcode: bool) -> None:
        super().__init__(serializer_data, endpoint_pref)
        self.by_barcode = by_barcode
        self.include = serializer_data.get('include') if self.by_barcode else 'all'
        self.less_than = serializer_data.get('less_than') if self.by_barcode else 1
        self.group_by = 'by_barcode' if self.by_barcode else serializer_data.get('group_by')
        self.is_excel = serializer_data.get('excel')

    def _scan_products_generator(self):
        zones_filter_params = {
            'status': ZoneStatusChoices.READY,
        }

        tasks_filter_params = {
            'status__in': (TaskStatusChoices.SUCCESS_SCAN, TaskStatusChoices.WORKED),
            'type': TaskTypeChoices.COUNTER_SCAN,
            'counter_scan_document__status': DocumentStatusChoices.READY,
        }

        zones = GetZonesService(
            project=self.project,
            zones_filter_params=zones_filter_params,
            tasks_filter_params=tasks_filter_params,
        ).process()

        for zone in zones:
            for task in zone.tasks.all():
                for scan_prd in task.scanned_products.all():  # noqa: WPS526
                    yield zone, scan_prd

    def _get_context(self) -> dict:  # noqa: WPS231
        source_data = {}
        # Собрать сырые данные, товары и зоны в которых они встречаются
        for zone, scan_prd in self._scan_products_generator():
            # Если выбран параметр - исключить зоны без названия склада
            if self.include == 'in_warehouses' and not zone.storage_name:
                continue

            if self.by_barcode:
                uniq_id = scan_prd.product.id
            else:
                uniq_id = scan_prd.dm if scan_prd.dm else scan_prd.product.dm

            if not uniq_id:
                continue

            if uniq_id not in source_data:
                code = scan_prd.product.barcode
                if self.group_by == 'by_product_code':
                    code = scan_prd.product.vendor_code
                dm = scan_prd.dm if scan_prd.dm else scan_prd.product.dm
                if not self.is_excel and dm:
                    # Разделить dm пробелами для pdf
                    split_dm = ''
                    counter = 0
                    for i in range(0, len(dm), 15):
                        split_dm += dm[counter:i] + ' '
                        counter = i
                    dm = split_dm
                source_data[uniq_id] = {
                    'barcode': scan_prd.product.barcode,
                    'dm': dm,
                    'title': MergeProductTitleAndTitleAttrsService(
                        additional_title_attrs=scan_prd.product.additional_title_attrs,
                        hidden_title_attrs=scan_prd.product.hide_title_attrs,
                        product_title=scan_prd.product.title,
                    ).process(),
                    'vendor_code': scan_prd.product.vendor_code,
                    'code': code,
                    'zones': {},
                }
            if zone.serial_number not in source_data[uniq_id]['zones']:
                source_data[uniq_id]['zones'][zone.serial_number] = float(scan_prd.amount)
            else:
                source_data[uniq_id]['zones'][zone.serial_number] += float(scan_prd.amount)
        zones_pairs = {'zones_pairs': {}}
        for _, product in source_data.items():
            zones = list(product.get('zones').items())
            for i in range(0, len(zones) - 1):
                for j in range(i + 1, len(zones)):  # noqa: WPS518
                    if zones[i][1] == zones[j][1]:
                        zones_pair_str = f'{zones[i][0]}/{zones[j][0]}'
                        if zones_pair_str not in zones_pairs['zones_pairs']:
                            zones_pairs['zones_pairs'][zones_pair_str] = [{
                                'barcode': product['barcode'],
                                'dm': product.get('dm'),
                                'title': product['title'],
                                'vendor_code': product['vendor_code'],
                                'code': product['code'],
                                'amount': zones[j][1],
                            }]
                        else:
                            zones_pairs['zones_pairs'][zones_pair_str].append({
                                'barcode': product['barcode'],
                                'dm': product.get('dm'),
                                'title': product['title'],
                                'vendor_code': product['vendor_code'],
                                'code': product['code'],
                                'amount': zones[j][1],
                            })

        less_than_pairs = {'zones_pairs': {}}
        for zone_pair, products in zones_pairs['zones_pairs'].items():
            if len(products) >= self.less_than:
                less_than_pairs['zones_pairs'][zone_pair] = products
        less_than_pairs['zones_pairs'] = dict(
            sorted(less_than_pairs['zones_pairs'].items(), key=lambda x: int(x[0].split('/')[0])),
        )
        for pair, _ in less_than_pairs['zones_pairs'].items():
            less_than_pairs['zones_pairs'][pair] = sorted(less_than_pairs['zones_pairs'][pair], key=lambda x: x['code'])
        return less_than_pairs

    def _generate_excel(self, source_data):  # noqa: WPS213 WPS231
        self.workbook = load_workbook(filename='api/v1/reports/excels/BarcodeMatches.xlsx')
        if not source_data:
            return
        worksheet = self.workbook.worksheets[0]
        times_new_roman_font = Font(name='Times New Roman', size=11)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
        )
        fill = PatternFill(start_color='DCDCDC', end_color='DCDCDC', fill_type='solid')
        if self.by_barcode:
            worksheet.cell(row=1, column=1, value='Совпадения ШК по зонам')
        else:
            worksheet.cell(row=1, column=1, value='Совпадения DM по зонам')
        row_pointer = 3
        for key, value in source_data.get('zones_pairs').items():
            worksheet.merge_cells(f'B{row_pointer}:C{row_pointer}')
            worksheet.cell(row=row_pointer, column=2, value='Совпадения в зонах')
            worksheet.cell(row=row_pointer, column=2).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=2).border = border
            worksheet.cell(row=row_pointer, column=2).fill = fill
            worksheet.cell(row=row_pointer, column=3).border = border
            worksheet.cell(row=row_pointer, column=4).border = border

            worksheet.merge_cells(f'E{row_pointer}:H{row_pointer}')
            worksheet.cell(row=row_pointer, column=5, value=key)
            worksheet.cell(row=row_pointer, column=5).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=5).border = border
            worksheet.cell(row=row_pointer, column=6).border = border
            worksheet.cell(row=row_pointer, column=7).border = border
            worksheet.cell(row=row_pointer, column=8).border = border

            row_pointer += 1
            worksheet.cell(row=row_pointer, column=2, value='№ п/п')
            worksheet.cell(row=row_pointer, column=2).fill = fill
            worksheet.cell(row=row_pointer, column=2).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=2).border = border
            worksheet.merge_cells(f'C{row_pointer}:E{row_pointer}')
            worksheet.cell(row=row_pointer, column=3, value='наименование')
            worksheet.cell(row=row_pointer, column=3).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=3).border = border
            worksheet.cell(row=row_pointer, column=3).fill = fill
            worksheet.cell(row=row_pointer, column=4).border = border
            worksheet.cell(row=row_pointer, column=5).border = border
            if self.by_barcode:
                worksheet.cell(row=row_pointer, column=6, value='ШК')
            else:
                worksheet.cell(row=row_pointer, column=6, value='DM')
            worksheet.cell(row=row_pointer, column=6).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=6).border = border
            worksheet.cell(row=row_pointer, column=6).fill = fill
            worksheet.cell(row=row_pointer, column=7, value='код товара')
            worksheet.cell(row=row_pointer, column=7).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=7).border = border
            worksheet.cell(row=row_pointer, column=7).fill = fill
            worksheet.cell(row=row_pointer, column=8, value='количество')
            worksheet.cell(row=row_pointer, column=8).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=8).border = border
            worksheet.cell(row=row_pointer, column=8).fill = fill

            row_pointer += 1
            for idx, product in enumerate(value, start=1):
                worksheet.cell(row=row_pointer, column=2, value=idx)
                worksheet.cell(row=row_pointer, column=2).font = times_new_roman_font
                worksheet.cell(row=row_pointer, column=2).border = border
                worksheet.merge_cells(f'C{row_pointer}:E{row_pointer}')
                worksheet.cell(row=row_pointer, column=3, value=product.get('title'))
                worksheet.cell(row=row_pointer, column=3).font = times_new_roman_font
                worksheet.cell(row=row_pointer, column=3).border = border
                worksheet.cell(row=row_pointer, column=4).border = border
                worksheet.cell(row=row_pointer, column=5).border = border
                if self.by_barcode:
                    worksheet.cell(row=row_pointer, column=6, value=product.get('barcode'))
                else:
                    worksheet.cell(
                        row=row_pointer,
                        column=6,
                        value=''.join([escape_xlsx_char(ch) for ch in product.get('dm')]),
                    )
                worksheet.cell(row=row_pointer, column=6).font = times_new_roman_font
                worksheet.cell(row=row_pointer, column=6).border = border
                if self.by_barcode:
                    worksheet.cell(row=row_pointer, column=7, value=product.get('vendor_code'))
                else:
                    worksheet.cell(row=row_pointer, column=7, value=product.get('code'))
                worksheet.cell(row=row_pointer, column=7).font = times_new_roman_font
                worksheet.cell(row=row_pointer, column=7).border = border
                worksheet.cell(row=row_pointer, column=8, value=str(product.get('amount')))
                worksheet.cell(row=row_pointer, column=8).font = times_new_roman_font
                worksheet.cell(row=row_pointer, column=8).border = border
                row_pointer += 1

            row_pointer += 2

    def process(self, *args, **kwargs):
        date = check_reports_directory()
        context = self._get_context()
        project_title = self.project.title.replace('/', '_').replace(' ', '_')
        if self.by_barcode:
            filename = f'Совпадения_ШК_по_зонам_{project_title}_{token_urlsafe(5)}.{"xlsx" if self.is_excel else "pdf"}'
        else:
            filename = f'Совпадения_DM_по_зонам_{project_title}_{token_urlsafe(5)}.{"xlsx" if self.is_excel else "pdf"}'
        if self.is_excel:
            self._generate_excel(context)
            self.workbook.save(f'{settings.MEDIA_ROOT}/reports/{date}/{filename}')
        else:
            if self.by_barcode:
                template_path = 'reports/barcodes-matches-report.html'
            else:
                template_path = 'reports/data-matrix-matches-report.html'
            res = render_to_string(template_path, context=context)
            output_dest = f'{settings.MEDIA_ROOT}/reports/{date}/{filename}'
            with open(output_dest, 'wb+') as ofile:
                pisa.CreatePDF(res, ofile)

        return f'{self.endpoint_pref}/media/reports/{date}/{filename}'
