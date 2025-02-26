import logging
from decimal import Decimal
from secrets import token_urlsafe

from django.conf import settings
from django.template.loader import render_to_string
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from xhtml2pdf import pisa

from api.v1.document.services.get_documents_for_reports import GetDocumentsForReportsService
from api.v1.product.services import GetProductsForReportsService, MergeProductTitleAndTitleAttrsService
from api.v1.reports.services.base_service import BaseReportService
from api.v1.reports.services.helpers.check_reports_directory import check_reports_directory
from apps import app
from apps.document.models import DocumentStatusChoices

logger = logging.getLogger('django')


@app.task
def generate_inv_ninteen_report_task(serializer_data: dict, endpoint_prefx: str):
    return InvNinteenReportService(serializer_data, endpoint_prefx).process()


class InvNinteenReportService(BaseReportService):  # noqa: WPS338
    """Сервис для генерации отчета ИНВ-19."""

    def __init__(self, serializer_data: dict, endpoint_pref: str) -> None:
        super().__init__(serializer_data, endpoint_pref)
        self.group_by = serializer_data['group_by']
        self.include = serializer_data['include']
        self.is_excel = serializer_data.get('excel')
        self.workbook = None

    def _chunks(self, lst, n):
        """Получение последовательных фрагментов размером n из lst."""  # noqa: DAR301
        for i in range(0, len(lst), n):  # noqa: WPS526
            yield lst[i:i + n]

    def _create_context(self):  # noqa: WPS231
        products = {}

        project_products = GetProductsForReportsService(project=self.project).process()

        for product in project_products:
            if (  # noqa: WPS337
                self.include == 'all'  # noqa: WPS222
                or self.include == 'only_identified' and product.title != 'Неизвестный товар'
                or self.include == 'only_unidentified' and product.title == 'Неизвестный товар'
            ):
                code = product.barcode if self.group_by == 'by_barcode' else product.vendor_code
                title_attrs = product.additional_title_attrs
                hide_attrs = product.hide_title_attrs
                products[code] = {
                    'code': code,
                    'title': MergeProductTitleAndTitleAttrsService(
                        additional_title_attrs=title_attrs,
                        hidden_title_attrs=hide_attrs,
                        product_title=product.title,
                    ).process(),
                    'price': product.price,
                    'remainder': product.remainder,
                }

        documents_filter_params = {
            'zone__project': self.project,
            'status': DocumentStatusChoices.READY,
        }

        documents = GetDocumentsForReportsService(documents_filter_params=documents_filter_params).process()

        for document in documents:
            for scanned_product in document.counter_scan_task.scanned_products.all():
                if (  # noqa: WPS337
                    self.include == 'all'  # noqa: WPS222
                    or self.include == 'found_by_product_code' and scanned_product.added_by_product_code is True
                    or self.include != 'found_by_product_code' and scanned_product.added_by_product_code is False
                ):
                    code = scanned_product.product.barcode
                    if self.group_by == 'by_product_code':
                        code = scanned_product.product.vendor_code
                    if self.include == 'found_by_product_code' and scanned_product.added_by_product_code is True:
                        title_attrs = scanned_product.product.additional_title_attrs
                        hide_attrs = scanned_product.product.hide_title_attrs
                        products[code] = {
                            'code': code,
                            'title': MergeProductTitleAndTitleAttrsService(
                                additional_title_attrs=title_attrs,
                                hidden_title_attrs=hide_attrs,
                                product_title=scanned_product.product.title,
                            ).process(),
                            'price': scanned_product.product.price,
                            'remainder': scanned_product.product.remainder,
                        }
                    if code not in products:
                        continue
                    if 'amount' not in products[code]:
                        products[code]['amount'] = scanned_product.amount
                    else:
                        products[code]['amount'] += scanned_product.amount

        for product in products.keys():
            if 'amount' not in products[product]:
                products[product]['amount'] = Decimal('0.000')
        products = [item for item in products.values() if item['remainder'] != item['amount']]

        alls_surplus_count = 0  # общее количество излишек
        alls_surplus_sum = 0  # общая сумма излишек
        alls_shortage_count = 0  # общее количество недостач
        alls_shortage_sum = 0  # общая сумма недостач

        corr_products = []
        # Считаю для каждого товара излишек или недостачу, в список только с разницей != 0
        for product in products:
            remainder = product.get('remainder')
            amount = product.get('amount')
            # Если разница между amount (факт кол-во) и учетным количеством положительна,
            # то это излишек. Если отрицательна - то недостача. Если ноль - не пускаем в отчет.
            # Может быть отрицательный остаток - в таком случае вычитаем из остатка отсканированное, это излишек с +.
            product['surplus'] = None
            product['surplus_sum'] = None
            product['shortage'] = None
            product['shortage_sum'] = None

            if remainder < 0:
                difference = abs(remainder - amount)
            else:
                difference = amount - remainder

            if difference == 0:
                continue
            if remainder < 0 or (remainder >= 0 and difference > 0):
                product['surplus'] = round(abs(difference), 2)
                product['surplus_sum'] = round(abs(difference) * product.get('price'), 2)
            elif remainder > 0 and difference < 0:  # noqa: WPS333
                product['shortage'] = round(abs(difference), 2)
                product['shortage_sum'] = round(abs(difference) * product.get('price'), 2)

            product['title'] = product.get('title').replace('_', ' ')
            if len(product.get('code')) > 13:
                product['code'] = product.get('code')[:14] + ' ' + product.get('code')[14:]

            if product['surplus']:
                alls_surplus_count += product['surplus']
            if product['surplus_sum']:
                alls_surplus_sum += product['surplus_sum']
            if product['shortage']:
                alls_shortage_count += product['shortage']
            if product['shortage_sum']:
                alls_shortage_sum += product['shortage_sum']

            corr_products.append(product)

        corr_products = sorted(corr_products, key=lambda x: x['code'])
        for idx, product in enumerate(corr_products, start=1):
            product['idx'] = idx

        context = {'products': corr_products}
        context['alls_surplus_count'] = alls_surplus_count
        if alls_surplus_count:
            context['alls_surplus_count'] = round(alls_surplus_count, 2)
        context['alls_surplus_sum'] = alls_surplus_sum
        if alls_surplus_sum:
            context['alls_surplus_sum'] = round(alls_surplus_sum, 2)
        context['alls_shortage_count'] = alls_shortage_count
        if alls_shortage_count:
            context['alls_shortage_count'] = round(alls_shortage_count, 2)
        context['alls_shortage_sum'] = alls_shortage_sum
        if alls_shortage_sum:
            context['alls_shortage_sum'] = round(alls_shortage_sum, 2)

        return context

    def _generate_excel(self, source_data):  # noqa: WPS213
        self.workbook = load_workbook(filename='api/v1/reports/excels/INV-19.xlsx')
        times_new_roman_font = Font(name='Times New Roman', size=10)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
        )
        top_medium_border = Border(top=Side(style='medium', color='000000'))
        alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        worksheet = self.workbook.worksheets[1]
        row_pointer = 6
        # В цикле проходим с первой ячейки 6 строки, подставляем необходимые данные из source_data
        for product_data in source_data.get('products'):
            worksheet.cell(row=row_pointer, column=1, value=product_data['idx'])
            worksheet.cell(row=row_pointer, column=1).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=1).border = border
            worksheet.cell(row=row_pointer, column=1).alignment = alignment

            worksheet.cell(row=row_pointer, column=2, value=product_data['title'])
            worksheet.cell(row=row_pointer, column=2).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=2).border = border
            worksheet.cell(row=row_pointer, column=2).alignment = alignment

            worksheet.cell(row=row_pointer, column=3, value=product_data['code'])
            worksheet.cell(row=row_pointer, column=3).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=3).border = border
            worksheet.cell(row=row_pointer, column=3).alignment = alignment

            worksheet.cell(row=row_pointer, column=4).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=4).border = border
            worksheet.cell(row=row_pointer, column=5).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=5).border = border
            worksheet.cell(row=row_pointer, column=6).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=6).border = border
            worksheet.cell(row=row_pointer, column=7).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=7).border = border

            worksheet.cell(row=row_pointer, column=8, value=product_data['surplus'])
            worksheet.cell(row=row_pointer, column=8).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=8).border = border
            worksheet.cell(row=row_pointer, column=8).number_format = '#,##0.00'
            worksheet.cell(row=row_pointer, column=8).alignment = alignment

            worksheet.cell(row=row_pointer, column=9, value=product_data['surplus_sum'])
            worksheet.cell(row=row_pointer, column=9).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=9).border = border
            worksheet.cell(row=row_pointer, column=9).number_format = '#,##0.00'
            worksheet.cell(row=row_pointer, column=9).alignment = alignment

            worksheet.cell(row=row_pointer, column=10, value=product_data['shortage'])
            worksheet.cell(row=row_pointer, column=10).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=10).border = border
            worksheet.cell(row=row_pointer, column=10).number_format = '#,##0.00'
            worksheet.cell(row=row_pointer, column=10).alignment = alignment

            worksheet.cell(row=row_pointer, column=11, value=product_data['shortage_sum'])
            worksheet.cell(row=row_pointer, column=11).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=11).border = border
            worksheet.cell(row=row_pointer, column=11).number_format = '#,##0.00'
            worksheet.cell(row=row_pointer, column=11).alignment = alignment

            worksheet.cell(row=row_pointer, column=12).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=12).border = border
            worksheet.cell(row=row_pointer, column=13).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=13).border = border
            worksheet.cell(row=row_pointer, column=14).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=14).border = border
            worksheet.cell(row=row_pointer, column=15).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=15).border = border
            worksheet.cell(row=row_pointer, column=16).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=16).border = border
            worksheet.cell(row=row_pointer, column=17).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=17).border = border

            worksheet.row_dimensions[row_pointer].height = 45

            row_pointer += 1

        worksheet.cell(row=row_pointer, column=7, value='Итого:')
        worksheet.cell(row=row_pointer, column=7).font = times_new_roman_font
        worksheet.cell(row=row_pointer, column=7).border = border
        worksheet.row_dimensions[row_pointer].height = 44

        worksheet.cell(row=row_pointer, column=8, value=source_data['alls_surplus_count'])
        worksheet.cell(row=row_pointer, column=8).font = times_new_roman_font
        worksheet.cell(row=row_pointer, column=8).border = border
        worksheet.cell(row=row_pointer, column=8).number_format = '#,##0.00'

        worksheet.cell(row=row_pointer, column=9, value=source_data['alls_surplus_sum'])
        worksheet.cell(row=row_pointer, column=9).font = times_new_roman_font
        worksheet.cell(row=row_pointer, column=9).border = border
        worksheet.cell(row=row_pointer, column=9).number_format = '#,##0.00'

        worksheet.cell(row=row_pointer, column=10, value=source_data['alls_shortage_count'])
        worksheet.cell(row=row_pointer, column=10).font = times_new_roman_font
        worksheet.cell(row=row_pointer, column=10).border = border
        worksheet.cell(row=row_pointer, column=10).number_format = '#,##0.00'

        worksheet.cell(row=row_pointer, column=11, value=source_data['alls_shortage_sum'])
        worksheet.cell(row=row_pointer, column=11).font = times_new_roman_font
        worksheet.cell(row=row_pointer, column=11).border = border
        worksheet.cell(row=row_pointer, column=11).number_format = '#,##0.00'

        row_pointer += 3

        worksheet.merge_cells(f'A{row_pointer}:B{row_pointer}')
        worksheet.cell(row=row_pointer, column=1, value='Бугалтер')
        worksheet.cell(row=row_pointer, column=1).font = times_new_roman_font

        worksheet.merge_cells(f'C{row_pointer}:F{row_pointer}')
        worksheet.merge_cells(f'H{row_pointer}:M{row_pointer}')
        row_pointer += 1
        worksheet.merge_cells(f'C{row_pointer}:F{row_pointer}')
        worksheet.merge_cells(f'H{row_pointer}:M{row_pointer}')
        worksheet.cell(row=row_pointer, column=3).border = top_medium_border
        worksheet.cell(row=row_pointer, column=4).border = top_medium_border
        worksheet.cell(row=row_pointer, column=5).border = top_medium_border
        worksheet.cell(row=row_pointer, column=6).border = top_medium_border
        worksheet.cell(row=row_pointer, column=3, value='подпись')
        worksheet.cell(row=row_pointer, column=3).font = times_new_roman_font
        worksheet.cell(row=row_pointer, column=8).border = top_medium_border
        worksheet.cell(row=row_pointer, column=9).border = top_medium_border
        worksheet.cell(row=row_pointer, column=10).border = top_medium_border
        worksheet.cell(row=row_pointer, column=11).border = top_medium_border
        worksheet.cell(row=row_pointer, column=12).border = top_medium_border
        worksheet.cell(row=row_pointer, column=13).border = top_medium_border
        worksheet.cell(row=row_pointer, column=8, value='расшифровка подписи')
        worksheet.cell(row=row_pointer, column=8).font = times_new_roman_font

    def process(self, *args, **kwargs):
        date = check_reports_directory()
        context = self._create_context()
        project_title = self.project.title.replace(' ', '_').replace('/', '_')
        filename = f'ИНВ-19_{project_title}_{token_urlsafe(5)}.{"xlsx" if self.is_excel else "pdf"}'

        if self.is_excel:
            self._generate_excel(context)
            self.workbook.save(f'{settings.MEDIA_ROOT}/reports/{date}/{filename}')
        else:
            res = render_to_string('reports/inv-19.html', context=context)
            output_dest = f'{settings.MEDIA_ROOT}/reports/{date}/{filename}'
            with open(output_dest, 'wb+') as ofile:
                pisa.CreatePDF(res, ofile)

        return f'{self.endpoint_pref}/media/reports/{date}/{filename}'
