from datetime import datetime
from secrets import token_urlsafe

from django.conf import settings
from django.template.loader import render_to_string
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from xhtml2pdf import pisa

from api.v1.reports.services.base_service import BaseReportService
from api.v1.reports.services.helpers.check_reports_directory import check_reports_directory
from apps import app
from apps.document.models import Document


@app.task
def generate_errors_report_task(serializer_data: dict, endpoint_prefix: str):
    """Таска для генерации отчета по ошибкам."""
    return ErrorsReportService(serializer_data, endpoint_prefix).process()


class ErrorsReportService(BaseReportService):  # noqa: WPS338
    """Сервис для отчета по ошибкам."""

    def __init__(self, serializer_data: dict, endpoint_pref: str):
        super().__init__(serializer_data, endpoint_pref)
        self.is_excel = serializer_data.get('excel')

    def _create_context_for_report(self):  # noqa: WPS231
        source_data = {}

        for document in Document.objects.filter(zone__project=self.project):
            if document.counter_scan_task:
                diff = False
                # Даже если С.УК не присылал данные, считаем, что данные не сошлись
                if (  # noqa: WPS337
                    document.controller_task
                    and document.counter_scan_task.result != document.controller_task.result
                ):
                    diff = True
                elif not document.controller_task:
                    diff = True

                if diff:
                    if document.counter_scan_task.zone not in source_data:
                        source_data[document.counter_scan_task.zone] = {
                            'zone': document.counter_scan_task.zone.serial_number,
                            'storage_name': document.counter_scan_task.zone.storage_name,
                            'scan': document.counter_scan_task.result,
                            'uk': document.controller_task.result if document.controller_task else ' ',
                        }

        clean_source_data = {'date': datetime.now(), 'storages': {}}
        for zone_data in source_data.values():
            if zone_data['storage_name'] not in clean_source_data['storages']:
                clean_source_data['storages'][zone_data['storage_name']] = []
            clean_source_data['storages'][zone_data['storage_name']].append(zone_data)

        return clean_source_data

    def _generate_excel(self, context):  # noqa: WPS213
        worksheet = self.workbook.worksheets[0]
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        times_new_roman_font = Font(name='Times New Roman', size=14)

        worksheet.cell(row=3, column=1, value=f'Дата: {datetime.now().strftime("%d-%m-%Y")}')
        worksheet.cell(row=3, column=1).font = times_new_roman_font

        row_pointer = 5
        for storage, zones_data in context['storages'].items():
            worksheet.cell(row=row_pointer, column=1, value=f'Название склада: {storage}')
            worksheet.cell(row=row_pointer, column=1).font = times_new_roman_font

            row_pointer += 1

            worksheet.cell(row=row_pointer, column=1, value='Номер зоны')
            worksheet.cell(row=row_pointer, column=1).border = thin_border
            worksheet.cell(row=row_pointer, column=1).alignment = Alignment(horizontal='left')
            worksheet.cell(row=row_pointer, column=1).font = times_new_roman_font

            worksheet.cell(row=row_pointer, column=2, value='Количество Скан')
            worksheet.cell(row=row_pointer, column=2).border = thin_border
            worksheet.cell(row=row_pointer, column=2).alignment = Alignment(horizontal='left')
            worksheet.cell(row=row_pointer, column=2).font = times_new_roman_font

            worksheet.cell(row=row_pointer, column=3, value='Количество УК')
            worksheet.cell(row=row_pointer, column=3).border = thin_border
            worksheet.cell(row=row_pointer, column=3).alignment = Alignment(horizontal='left')
            worksheet.cell(row=row_pointer, column=3).font = times_new_roman_font

            worksheet.cell(row=row_pointer, column=4, value='Номер зоны для проверки')
            worksheet.cell(row=row_pointer, column=4).border = thin_border
            worksheet.cell(row=row_pointer, column=4).alignment = Alignment(horizontal='left')
            worksheet.cell(row=row_pointer, column=4).font = times_new_roman_font

            row_pointer += 1

            for zone in zones_data:
                worksheet.cell(row=row_pointer, column=1, value=zone['zone'])
                worksheet.cell(row=row_pointer, column=1).border = thin_border
                worksheet.cell(row=row_pointer, column=1).alignment = Alignment(horizontal='left')
                worksheet.cell(row=row_pointer, column=1).font = times_new_roman_font

                worksheet.cell(row=row_pointer, column=2, value=zone['scan'])
                worksheet.cell(row=row_pointer, column=2).border = thin_border
                worksheet.cell(row=row_pointer, column=2).alignment = Alignment(horizontal='left')
                worksheet.cell(row=row_pointer, column=2).font = times_new_roman_font

                worksheet.cell(row=row_pointer, column=3, value=zone['uk'])
                worksheet.cell(row=row_pointer, column=3).border = thin_border
                worksheet.cell(row=row_pointer, column=3).alignment = Alignment(horizontal='left')
                worksheet.cell(row=row_pointer, column=3).font = times_new_roman_font

                worksheet.cell(row=row_pointer, column=4, value=f'{zone["zone"]} ____________________')
                worksheet.cell(row=row_pointer, column=4).border = thin_border
                worksheet.cell(row=row_pointer, column=4).alignment = Alignment(horizontal='left')
                worksheet.cell(row=row_pointer, column=4).font = times_new_roman_font
                row_pointer += 1

            row_pointer += 2

    def process(self, *args, **kwargs):
        date = check_reports_directory()
        context = self._create_context_for_report()
        filename = f'errors-report_{token_urlsafe(5)}.{"xlsx" if self.is_excel else "pdf"}'

        if self.is_excel:
            self.workbook = load_workbook(filename='api/v1/reports/excels/errors.xlsx')
            self._generate_excel(context=context)
            self.workbook.save(f'{settings.MEDIA_ROOT}/reports/{date}/{filename}')
        else:
            res = render_to_string('reports/errors-report.html', context=context)
            output_dest = f'{settings.MEDIA_ROOT}/reports/{date}/{filename}'
            with open(output_dest, 'wb+') as ofile:
                pisa.CreatePDF(res, ofile)

        return f'{self.endpoint_pref}/media/reports/{date}/{filename}'
