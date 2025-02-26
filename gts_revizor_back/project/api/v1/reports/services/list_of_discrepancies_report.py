import logging
from decimal import Decimal
from secrets import token_urlsafe

from django.conf import settings
from django.template.loader import render_to_string
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from xhtml2pdf import pisa

from api.v1.document.services.get_documents_for_reports import GetDocumentsForReportsService
from api.v1.product.services import GetProductsForListOfDiscrepanciesReportsService
from api.v1.reports.services.base_service import BaseReportService
from api.v1.reports.services.helpers.check_reports_directory import check_reports_directory
from apps import app
from apps.document.models import DocumentStatusChoices

logger = logging.getLogger('django')


@app.task
def generate_list_of_discrepancies_report_task(serializer_data: dict, endpoint_prefx: str):
    return ListOfDiscrepanciesReportService(serializer_data, endpoint_prefx).process()


class ListOfDiscrepanciesReportService(BaseReportService):  # noqa: WPS338
    """Сервис для генерации отчета Ведомость расхождения Концепт Груп."""

    def __init__(self, serializer_data: dict, endpoint_pref: str):
        super().__init__(serializer_data, endpoint_pref)
        self.only_discrepancies = serializer_data.get('only_discrepancies')
        self.is_excel = serializer_data.get('excel')

    def _create_context(self, *args, **kwargs):  # noqa: WPS231
        source_data = {}

        products = GetProductsForListOfDiscrepanciesReportsService(project=self.project).process()

        for product in products:
            source_data[product] = {
                'project_id': product.project.id,
                'product_id': product.id,
                'barcode': product.barcode,
                'title': product.title,
                'vendor_code': product.vendor_code,
                'price': float(product.price),
                'remainder': product.remainder,
                'remainder_2': product.remainder_2,
                'size': product.size,
                'zone': [],
                'fact': 0,
            }

        documents_filter_params = {
            'zone__project': self.project,
            'status': DocumentStatusChoices.READY,
        }

        documents = GetDocumentsForReportsService(documents_filter_params=documents_filter_params).process()

        for document in documents.prefetch_related('counter_scan_task__scanned_products__task__zone'):
            if document.counter_scan_task is None:
                continue
            for scan_product in document.counter_scan_task.scanned_products.all():
                if scan_product.product not in source_data:
                    source_data[scan_product.product] = {
                        'barcode': scan_product.product.barcode,
                        'vendor_code': scan_product.product.vendor_code,
                        'title': scan_product.product.title,
                        'price': float(scan_product.product.price),
                        'zone': [scan_product.task.zone.serial_number],
                        'size': scan_product.product.size,
                        'remainder': scan_product.product.remainder if scan_product.product.remainder else 0,
                        'remainder_2': scan_product.product.remainder_2 if scan_product.product.remainder_2 else 0,
                        'fact': scan_product.amount,
                    }
                    source_data[scan_product.product]['zone'] = [scan_product.task.zone.serial_number]
                    source_data[scan_product.product]['fact'] = scan_product.amount
                else:
                    source_data[scan_product.product]['fact'] += scan_product.amount
                    zone_ser_number = scan_product.task.zone.serial_number
                    if zone_ser_number not in source_data[scan_product.product]['zone']:
                        source_data[scan_product.product]['zone'].append(zone_ser_number)

        context = {
            'project': self.project.title,
            'products': [],
            'alls_fact_count': 0,
            'alls_diff_count': 0,
            'alls_diff_price': 0,
        }

        for product in source_data.keys():
            remainder = round(source_data[product]['remainder'], 3)
            remainder_2 = round(source_data[product]['remainder_2'], 3)
            fact = round(source_data[product]['fact'], 3)
            diff_count = round(fact - (remainder + remainder_2), 3)
            total_remainder = round(remainder + remainder_2, 3)
            diff_price = round(diff_count * Decimal(source_data[product]['price']), 2)

            if remainder == 0 and remainder_2 == 0 and fact == 0:
                continue

            if self.only_discrepancies:
                if diff_count == 0:
                    continue

            product = source_data[product]
            zones = ', '.join([str(zone_num) for zone_num in product['zone']])
            context['products'].append({
                'barcode': product['barcode'],
                'vendor_code': product['vendor_code'],
                'title': product['title'],
                'price': round(product['price'], 2),
                'size': product['size'] or '',
                'zone': zones,
                'remainder': remainder,
                'remainder_2': remainder_2,
                'total_remainder': total_remainder,
                'fact': fact,
                'diff_count': diff_count,
                'diff_price': diff_price,
            })

            context['alls_fact_count'] += product['fact']
            context['alls_diff_count'] += diff_count
            context['alls_diff_price'] += diff_price

        context['alls_fact_count'] = round(context['alls_fact_count'], 3)
        context['alls_diff_count'] = round(context['alls_diff_count'], 3)
        context['alls_diff_price'] = round(context['alls_diff_price'], 2)

        return context

    def _generate_excel(self, context):  # noqa: WPS213
        worksheet = self.workbook.worksheets[0]

        times_new_roman_font = Font(name='Times New Roman', size=9)
        alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        row_pointer = 4
        for idx, product in enumerate(context.get('products'), start=1):
            worksheet.cell(row=row_pointer, column=1, value=idx)
            worksheet.cell(row=row_pointer, column=1).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=1).alignment = alignment

            worksheet.cell(row=row_pointer, column=2, value=product['barcode'])
            worksheet.cell(row=row_pointer, column=2).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=2).alignment = alignment

            worksheet.cell(row=row_pointer, column=3, value=product['vendor_code'])
            worksheet.cell(row=row_pointer, column=3).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=3).alignment = alignment

            worksheet.cell(row=row_pointer, column=4, value=product['size'])
            worksheet.cell(row=row_pointer, column=4).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=4).alignment = alignment

            worksheet.cell(row=row_pointer, column=5, value=product['title'])
            worksheet.cell(row=row_pointer, column=5).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=5).alignment = alignment

            worksheet.cell(row=row_pointer, column=6, value=product['price'])
            worksheet.cell(row=row_pointer, column=6).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=6).alignment = alignment

            worksheet.cell(row=row_pointer, column=7, value=product['remainder'])
            worksheet.cell(row=row_pointer, column=7).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=7).alignment = alignment

            worksheet.cell(row=row_pointer, column=8, value=product['remainder_2'])
            worksheet.cell(row=row_pointer, column=8).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=8).alignment = alignment

            worksheet.cell(row=row_pointer, column=9, value=product['total_remainder'])
            worksheet.cell(row=row_pointer, column=9).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=9).alignment = alignment

            worksheet.cell(row=row_pointer, column=10, value=product['fact'])
            worksheet.cell(row=row_pointer, column=10).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=10).alignment = alignment

            worksheet.cell(row=row_pointer, column=11, value=product['diff_count'])
            worksheet.cell(row=row_pointer, column=11).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=11).alignment = alignment

            worksheet.cell(row=row_pointer, column=12, value=product['diff_price'])
            worksheet.cell(row=row_pointer, column=12).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=12).alignment = alignment

            worksheet.cell(row=row_pointer, column=13, value=product['zone'])
            worksheet.cell(row=row_pointer, column=13).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=13).alignment = alignment

            row_pointer += 1

        worksheet.merge_cells(f'A{row_pointer}:E{row_pointer}')
        worksheet.cell(row=row_pointer, column=1, value='Итого:')
        worksheet.cell(row=row_pointer, column=1).font = times_new_roman_font
        worksheet.cell(row=row_pointer, column=1).alignment = Alignment(horizontal='right', vertical='center')

        worksheet.cell(row=row_pointer, column=10, value=context['alls_fact_count'])
        worksheet.cell(row=row_pointer, column=10).font = times_new_roman_font
        worksheet.cell(row=row_pointer, column=10).alignment = alignment

        worksheet.cell(row=row_pointer, column=11, value=context['alls_diff_count'])
        worksheet.cell(row=row_pointer, column=11).font = times_new_roman_font
        worksheet.cell(row=row_pointer, column=11).alignment = alignment

        worksheet.cell(row=row_pointer, column=12, value=context['alls_diff_price'])
        worksheet.cell(row=row_pointer, column=12).font = times_new_roman_font
        worksheet.cell(row=row_pointer, column=12).alignment = alignment

    def process(self, *args, **kwargs):
        date = check_reports_directory()
        context = self._create_context()
        project_title = self.project.title.replace('/', '_').replace(' ', '_')
        filename = f'Ведомость_расхождений_{project_title}_{token_urlsafe(5)}.{"xlsx" if self.is_excel else "pdf"}'

        if self.is_excel:
            self.workbook = load_workbook(filename='api/v1/reports/excels/ListOfDiscrepancies.xlsx')
            self._generate_excel(context=context)
            self.workbook.save(f'{settings.MEDIA_ROOT}/reports/{date}/{filename}')
        else:
            res = render_to_string('reports/list-of-discrepancies.html', context=context)
            output_dest = f'{settings.MEDIA_ROOT}/reports/{date}/{filename}'
            with open(output_dest, 'wb+') as ofile:
                pisa.CreatePDF(res, ofile)

        return f'{self.endpoint_pref}/media/reports/{date}/{filename}'
