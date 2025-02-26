from secrets import token_urlsafe

from django.conf import settings
from django.template.loader import render_to_string
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from xhtml2pdf import pisa

from api.v1.reports.services.base_service import BaseReportService
from api.v1.reports.services.helpers.check_reports_directory import check_reports_directory
from apps import app
from apps.document.models import Document, DocumentStatusChoices


@app.task
def generate_inv_in_zones_report_task(serializer_data: dict, endpoint_prefx: str):
    return InvInZonesReportService(serializer_data, endpoint_prefx).process()


class InvInZonesReportService(BaseReportService):  # noqa: WPS338
    """Сервис для генерации отчета Общее кол-во ТМЦ по зонам."""

    def __init__(self, serializer_data: dict, endpoint_pref: str) -> None:
        super().__init__(serializer_data, endpoint_pref)
        self.serial_number_start = serializer_data.get('serial_number_start')
        self.serial_number_end = serializer_data.get('serial_number_end')
        self.is_excel = serializer_data.get('excel')

    def _create_context(self):
        documents = Document.objects.filter(
            zone__project=self.project,
            status=DocumentStatusChoices.READY,
            zone__serial_number__gte=self.serial_number_start,
            zone__serial_number__lte=self.serial_number_end,
        ).order_by('zone__serial_number')

        context = {'zones': []}
        for document in documents:
            context['zones'].append({
                'title': document.zone.title,
                'tmc_count': document.counter_scan_task.result,
            })

        return context

    def _generate_excel(self, context):
        worksheet = self.workbook.worksheets[0]

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        row_pointer = 4
        for idx, zone in enumerate(context['zones'], start=1):
            worksheet.cell(row=row_pointer, column=2, value=idx)
            worksheet.cell(row=row_pointer, column=3, value=zone.get('title'))
            worksheet.cell(row=row_pointer, column=4, value=zone.get('tmc_count'))
            for i in range(2, 5):
                worksheet.cell(row=row_pointer, column=i).border = thin_border
                worksheet.cell(row=row_pointer, column=i).alignment = Alignment(horizontal='left')
            row_pointer += 1

    def process(self, *args, **kwargs):
        date = check_reports_directory()
        context = self._create_context()
        project_title = self.project.title.replace(' ', '_').replace('/', '_')
        filename = (
            f'Общее_количество_ТМЦ_по_зонам_'
            f'{project_title}_{token_urlsafe(5)}.{"xlsx" if self.is_excel else "pdf"}'  # noqa: WPS326
        )

        if self.is_excel:
            self.workbook = load_workbook(filename='api/v1/reports/excels/InventoryInZones.xlsx')
            self._generate_excel(context=context)
            self.workbook.save(f'{settings.MEDIA_ROOT}/reports/{date}/{filename}')
        else:
            res = render_to_string('reports/inventory-in-zones.html', context=context)
            output_dest = f'{settings.MEDIA_ROOT}/reports/{date}/{filename}'
            with open(output_dest, 'wb+') as ofile:
                pisa.CreatePDF(res, ofile)

        return f'{self.endpoint_pref}/media/reports/{date}/{filename}'
