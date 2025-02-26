from secrets import token_urlsafe

from django.conf import settings
from django.template.loader import render_to_string
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from xhtml2pdf import pisa

from api.v1.reports.services.base_service import BaseReportService
from api.v1.reports.services.helpers.check_reports_directory import check_reports_directory
from apps import app
from apps.document.models import DocumentStatusChoices
from apps.zone.models import Zone, ZoneStatusChoices


@app.task
def inventory_in_zones_in_zones_report_celery_wrapper(serializer_data: dict, endpoint_pref: str):  # noqa: WPS118
    return InventoryInZonesInZonesReportService(serializer_data=serializer_data, endpoint_pref=endpoint_pref).process()


class InventoryInZonesInZonesReportService(BaseReportService):  # noqa: WPS338
    """Сервис для генерации отчета "Отчет данных инвентаризации по зонам" на экране Зоны."""

    def __init__(self, serializer_data: dict, endpoint_pref: str) -> None:
        super().__init__(serializer_data, endpoint_pref)
        self.document_type: str = serializer_data.get('document_type')
        self.group_by: str = serializer_data.get('group_by')
        self.zones: list = serializer_data.get('zones', None)

        self.workbook = None
        if self.document_type == 'excel':
            self.workbook = load_workbook(filename='api/v1/reports/excels/InventoryInZonesInZones.xlsx')

    def _create_context(self):  # noqa: WPS231
        zones = Zone.objects.filter(project=self.project, status=ZoneStatusChoices.READY)
        if self.zones:
            zones = Zone.objects.filter(pk__in=self.zones)

        context = {'zones': {}}
        for zone in zones:
            context['zones'][zone.serial_number] = {'products': [], 'total': 0}  # noqa: WPS204
            for document in zone.documents.filter(status=DocumentStatusChoices.READY):
                for product in document.counter_scan_task.scanned_products.all():
                    amount = product.amount
                    if float(amount) % 1 == 0.0:  # noqa: WPS459 WPS345 WPS358
                        amount = int(product.amount)

                    vendor_code = f'art_{product.product.barcode}'
                    if product.product.vendor_code:
                        vendor_code = product.product.vendor_code

                    context['zones'][zone.serial_number]['products'].append({
                        'barcode': product.product.barcode,
                        'vendor_code': vendor_code,
                        'title': product.product.title,
                        'amount': amount,
                    })
                    context['zones'][zone.serial_number]['total'] += amount

        clear_context = {'zones': {}}
        for zone in context['zones']:  # noqa: WPS528
            if context['zones'][zone].get('products'):
                clear_context['zones'][zone] = {
                    'products': context['zones'][zone].get('products'),
                    'total': context['zones'][zone].get('total'),
                }

        clear_context['zones'] = dict(sorted(clear_context['zones'].items()))
        for zone in clear_context['zones'].values():  # noqa: WPS426
            zone['products'] = sorted(zone['products'], key=lambda x: x[self.group_by])
        return clear_context

    def write_pdf(self, context, date):
        res = render_to_string('reports/inventory-in-zones-in-zones.html', context=context)
        project_title = self.project.title.replace(' ', '_').replace('/', '_')
        filename = f'Отчет_по_инвентаризации_в_зонах_{project_title}_{token_urlsafe(5)}.pdf'
        output_dest = f'{settings.MEDIA_ROOT}/reports/{date}/{filename}'
        with open(output_dest, 'wb+') as ofile:
            pisa.CreatePDF(res, ofile)
        return f'{self.endpoint_pref}/media/reports/{date}/{filename}'

    def write_excel(self, context, date):  # noqa: WPS213
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        times_new_roman_font = Font(name='Times New Roman', size=14)
        fill = PatternFill(
            start_color='DCDCDC',
            end_color='DCDCDC',
            fill_type='solid',
        )

        worksheet = self.workbook.worksheets[0]
        worksheet.cell(row=1, column=1).fill = fill
        row_pointer = 3

        for zone_number, data in context['zones'].items():
            worksheet.cell(row=row_pointer, column=1, value='Зона:')
            worksheet.cell(row=row_pointer, column=1).border = thin_border
            worksheet.cell(row=row_pointer, column=1).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=1).fill = fill

            worksheet.cell(row=row_pointer, column=2).fill = fill
            worksheet.merge_cells(f'B{row_pointer}:E{row_pointer}')
            worksheet.cell(row=row_pointer, column=2, value=zone_number).alignment = Alignment(horizontal='left')
            for i in range(2, 6):
                worksheet.cell(row=row_pointer, column=i).border = thin_border  # noqa: WPS204
                worksheet.cell(row=row_pointer, column=i).font = times_new_roman_font
                worksheet.cell(row=row_pointer, column=i).fill = fill
            row_pointer += 1

            worksheet.cell(row=row_pointer, column=1, value='Штрихкод')
            worksheet.cell(row=row_pointer, column=2, value='Артикул')
            worksheet.cell(row=row_pointer, column=3, value='Наименование')
            worksheet.cell(row=row_pointer, column=4, value='Кол-во')
            worksheet.cell(row=row_pointer, column=5, value='Внешнее контр. кол.')
            for i in range(1, 6):
                worksheet.cell(row=row_pointer, column=i).border = thin_border
                worksheet.cell(row=row_pointer, column=i).font = times_new_roman_font
            row_pointer += 1

            for product in data['products']:
                worksheet.cell(row=row_pointer, column=1, value=product['barcode'])
                worksheet.cell(row=row_pointer, column=2, value=product['vendor_code'])
                worksheet.cell(row=row_pointer, column=3, value=product['title'])
                worksheet.cell(row=row_pointer, column=4, value=product['amount'])
                worksheet.cell(row=row_pointer, column=5, value='')  # TODO

                for i in range(1, 6):
                    worksheet.cell(row=row_pointer, column=i).border = thin_border
                    worksheet.cell(row=row_pointer, column=i).font = times_new_roman_font

                row_pointer += 1

            worksheet.merge_cells(f'A{row_pointer}:C{row_pointer}')
            for i in range(1, 4):
                worksheet.cell(row=row_pointer, column=i).border = thin_border
            worksheet.cell(row=row_pointer, column=1, value='Всего: ').alignment = Alignment(horizontal='right')
            worksheet.cell(row=row_pointer, column=1).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=4, value=data.get('total'))
            worksheet.cell(row=row_pointer, column=4).font = times_new_roman_font
            worksheet.cell(row=row_pointer, column=4).border = thin_border
            worksheet.cell(row=row_pointer, column=5).border = thin_border

            row_pointer += 1
            worksheet.merge_cells(f'A{row_pointer}:E{row_pointer}')
            row_pointer += 1

        project_title = self.project.title.replace(' ', '_').replace('/', '_')
        filename = f'Отчет_по_инвентаризации_в_зонах_{project_title}_{token_urlsafe(5)}.xlsx'
        self.workbook.save(f'{settings.MEDIA_ROOT}/reports/{date}/{filename}')
        return f'{self.endpoint_pref}/media/reports/{date}/{filename}'

    def process(self, *args, **kwargs):
        date = check_reports_directory()
        context = self._create_context()

        if self.document_type == 'pdf':
            filepath = self.write_pdf(context=context, date=date)
        elif self.document_type == 'excel':
            filepath = self.write_excel(context=context, date=date)

        return filepath
